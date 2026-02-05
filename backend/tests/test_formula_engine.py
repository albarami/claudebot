"""
Unit tests for the Deterministic Formula Engine.
Tests the production engines/formula_engine.py module.
"""

import pytest
import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from engines.formula_engine import FormulaEngine
from models.task_schema import TaskType, TaskSpec


@pytest.fixture
def sample_df():
    """Create sample DataFrame for testing."""
    return pd.DataFrame({
        "ID": range(1, 101),
        "Age": [25 + i % 40 for i in range(100)],
        "Score": [50 + i % 50 for i in range(100)],
        "Group": ["A" if i % 2 == 0 else "B" for i in range(100)],
        "Rating": [3.5 + (i % 5) * 0.5 for i in range(100)]
    })


@pytest.fixture
def temp_workbook(sample_df):
    """Create a temporary workbook with raw data."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb = Workbook()
        ws = wb.active
        ws.title = "00_RAW_DATA_LOCKED"
        
        for col_idx, col_name in enumerate(sample_df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        
        for row_idx, row in enumerate(sample_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(f.name)
        yield Path(f.name)


class TestFormulaEngine:
    """Tests for the production FormulaEngine class."""
    
    def test_engine_initialization(self, temp_workbook, sample_df):
        """Test engine initializes with correct column mappings."""
        engine = FormulaEngine(
            workbook_path=temp_workbook,
            df=sample_df,
            session_id="test_session"
        )
        
        assert engine.n_rows == 100
        assert len(engine.col_mapping) == 5
        assert "Age" in engine.col_mapping
        assert engine.col_mapping["Age"] == "B"
    
    def test_numeric_column_detection(self, temp_workbook, sample_df):
        """Test engine correctly detects numeric columns."""
        engine = FormulaEngine(
            workbook_path=temp_workbook,
            df=sample_df,
            session_id="test_session"
        )
        
        assert "Age" in engine.numeric_cols
        assert "Score" in engine.numeric_cols
        assert "Rating" in engine.numeric_cols
        assert "Group" in engine.categorical_cols
    
    def test_col_mapping_letters(self, temp_workbook, sample_df):
        """Test column letter mapping is correct."""
        engine = FormulaEngine(
            workbook_path=temp_workbook,
            df=sample_df,
            session_id="test_session"
        )
        
        assert engine.col_mapping["ID"] == "A"
        assert engine.col_mapping["Age"] == "B"
        assert engine.col_mapping["Score"] == "C"
        assert engine.col_mapping["Group"] == "D"
        assert engine.col_mapping["Rating"] == "E"


class TestTaskExecution:
    """Tests for task execution methods."""
    
    def test_execute_descriptive_stats_task(self, temp_workbook, sample_df):
        """Test descriptive statistics task execution."""
        engine = FormulaEngine(
            workbook_path=temp_workbook,
            df=sample_df,
            session_id="test_session"
        )
        
        task = TaskSpec(
            id="1.1",
            name="Descriptive Statistics",
            task_type=TaskType.DESCRIPTIVE_STATS,
            phase="3_Descriptive",
            objective="Calculate descriptive statistics for all numeric variables",
            output_sheet="03_DESCRIPTIVES"
        )
        
        result = engine.execute_task(task)
        
        assert "sheet_name" in result
        assert "formulas" in result
        assert len(result["formulas"]) > 0
    
    def test_execute_missing_data_task(self, temp_workbook, sample_df):
        """Test missing data analysis task execution."""
        engine = FormulaEngine(
            workbook_path=temp_workbook,
            df=sample_df,
            session_id="test_session"
        )
        
        task = TaskSpec(
            id="1.2",
            name="Missing Data Analysis",
            task_type=TaskType.MISSING_DATA,
            phase="1_Data_Validation",
            objective="Analyze missing data patterns across all variables",
            output_sheet="01_MISSING"
        )
        
        result = engine.execute_task(task)
        
        assert "sheet_name" in result
        assert result["sheet_name"] == "01_MISSING"
    
    def test_execute_correlation_task(self, temp_workbook, sample_df):
        """Test correlation matrix task execution."""
        engine = FormulaEngine(
            workbook_path=temp_workbook,
            df=sample_df,
            session_id="test_session"
        )
        
        task = TaskSpec(
            id="2.1",
            name="Correlation Matrix",
            task_type=TaskType.CORRELATION_MATRIX,
            phase="4_Inferential",
            objective="Generate correlation matrix for numeric variables",
            output_sheet="04_CORRELATIONS"
        )
        
        result = engine.execute_task(task)
        
        assert "sheet_name" in result
        assert "formulas" in result


class TestFormulaGeneration:
    """Tests for formula string generation."""
    
    def test_format_criteria_string(self, temp_workbook, sample_df):
        """Test criteria formatting for string values."""
        engine = FormulaEngine(
            workbook_path=temp_workbook,
            df=sample_df,
            session_id="test_session"
        )
        
        criteria = engine._format_criteria("A")
        assert criteria == '"A"'
    
    def test_format_criteria_number(self, temp_workbook, sample_df):
        """Test criteria formatting for numeric values."""
        engine = FormulaEngine(
            workbook_path=temp_workbook,
            df=sample_df,
            session_id="test_session"
        )
        
        criteria = engine._format_criteria(42)
        assert criteria == "42"
    
    def test_format_criteria_boolean(self, temp_workbook, sample_df):
        """Test criteria formatting for boolean values."""
        engine = FormulaEngine(
            workbook_path=temp_workbook,
            df=sample_df,
            session_id="test_session"
        )
        
        assert engine._format_criteria(True) == "TRUE"
        assert engine._format_criteria(False) == "FALSE"
