"""
APA Reporting Module.
Generates APA 7th Edition formatted tables and interpretation text.
"""

from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass
from enum import Enum

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class EffectSizeInterpretation(str, Enum):
    """Cohen's conventions for effect size interpretation."""
    NEGLIGIBLE = "negligible"
    SMALL = "small"
    MEDIUM = "medium"
    LARGE = "large"


@dataclass
class APATableStyle:
    """APA 7 table formatting settings."""
    title_font: Font = None
    header_font: Font = None
    body_font: Font = None
    note_font: Font = None
    header_border: Border = None
    
    def __post_init__(self):
        self.title_font = Font(name='Times New Roman', size=12, bold=True)
        self.header_font = Font(name='Times New Roman', size=11, bold=True)
        self.body_font = Font(name='Times New Roman', size=11)
        self.note_font = Font(name='Times New Roman', size=10, italic=True)
        thin = Side(style='thin')
        self.header_border = Border(bottom=thin)


def interpret_cohens_d(d: float) -> EffectSizeInterpretation:
    """
    Interpret Cohen's d using standard conventions.
    
    Args:
        d: Cohen's d value (absolute).
    
    Returns:
        Effect size interpretation category.
    """
    d_abs = abs(d)
    if d_abs < 0.2:
        return EffectSizeInterpretation.NEGLIGIBLE
    elif d_abs < 0.5:
        return EffectSizeInterpretation.SMALL
    elif d_abs < 0.8:
        return EffectSizeInterpretation.MEDIUM
    else:
        return EffectSizeInterpretation.LARGE


def interpret_correlation(r: float) -> str:
    """
    Interpret Pearson correlation using standard conventions.
    
    Args:
        r: Correlation coefficient.
    
    Returns:
        Interpretation string.
    """
    r_abs = abs(r)
    if r_abs < 0.1:
        strength = "negligible"
    elif r_abs < 0.3:
        strength = "weak"
    elif r_abs < 0.5:
        strength = "moderate"
    elif r_abs < 0.7:
        strength = "strong"
    else:
        strength = "very strong"
    
    direction = "positive" if r >= 0 else "negative"
    return f"{strength} {direction}"


def interpret_cronbach_alpha(alpha: float) -> str:
    """
    Interpret Cronbach's alpha for internal consistency.
    
    Args:
        alpha: Cronbach's alpha value.
    
    Returns:
        Interpretation string.
    """
    if alpha >= 0.9:
        return "excellent"
    elif alpha >= 0.8:
        return "good"
    elif alpha >= 0.7:
        return "acceptable"
    elif alpha >= 0.6:
        return "questionable"
    elif alpha >= 0.5:
        return "poor"
    else:
        return "unacceptable"


def format_p_value(p: float) -> str:
    """
    Format p-value according to APA guidelines.
    
    Args:
        p: P-value.
    
    Returns:
        APA-formatted p-value string.
    """
    if p < 0.001:
        return "< .001"
    elif p < 0.01:
        return f"= {p:.3f}"[1:]  # Remove leading zero
    else:
        return f"= {p:.2f}"[1:]


def format_statistic(value: float, decimals: int = 2) -> str:
    """
    Format a statistic value for APA reporting.
    
    Args:
        value: Numeric value.
        decimals: Decimal places.
    
    Returns:
        Formatted string.
    """
    if abs(value) < 1:
        return f"{value:.{decimals}f}"[1:] if value >= 0 else f"-{abs(value):.{decimals}f}"[1:]
    return f"{value:.{decimals}f}"


class APATableWriter:
    """
    Writes APA 7-formatted tables to Excel worksheets.
    """
    
    def __init__(self, worksheet: Worksheet):
        self.ws = worksheet
        self.style = APATableStyle()
        self.current_row = 1
    
    def write_descriptives_table(
        self,
        variables: List[str],
        stats: Dict[str, Dict[str, float]],
        title: str = "Descriptive Statistics",
        start_row: int = 1
    ) -> int:
        """
        Write APA-formatted descriptive statistics table.
        
        Args:
            variables: List of variable names.
            stats: Dict mapping variable to {n, mean, sd, min, max, skew, kurt}.
            title: Table title.
            start_row: Starting row.
        
        Returns:
            Next available row after table.
        """
        row = start_row
        
        self.ws.cell(row=row, column=1, value=f"Table X")
        self.ws.cell(row=row, column=1).font = self.style.title_font
        row += 1
        
        self.ws.cell(row=row, column=1, value=title)
        self.ws.cell(row=row, column=1).font = Font(
            name='Times New Roman', size=12, italic=True
        )
        row += 2
        
        headers = ["Variable", "n", "M", "SD", "Min", "Max", "Skew", "Kurt"]
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=row, column=col, value=header)
            cell.font = self.style.header_font
            cell.border = self.style.header_border
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        for var in variables:
            var_stats = stats.get(var, {})
            self.ws.cell(row=row, column=1, value=var).font = self.style.body_font
            
            values = [
                var_stats.get('n', ''),
                f"={format_statistic(var_stats.get('mean', 0))}" if 'mean' in var_stats else '',
                f"={format_statistic(var_stats.get('sd', 0))}" if 'sd' in var_stats else '',
                var_stats.get('min', ''),
                var_stats.get('max', ''),
                f"={format_statistic(var_stats.get('skew', 0))}" if 'skew' in var_stats else '',
                f"={format_statistic(var_stats.get('kurt', 0))}" if 'kurt' in var_stats else ''
            ]
            
            for col, val in enumerate(values, 2):
                cell = self.ws.cell(row=row, column=col, value=val)
                cell.font = self.style.body_font
                cell.alignment = Alignment(horizontal='center')
            row += 1
        
        row += 1
        note = "Note. M = mean; SD = standard deviation; Skew = skewness; Kurt = kurtosis."
        self.ws.cell(row=row, column=1, value=note).font = self.style.note_font
        
        return row + 2

    def write_correlation_table(
        self,
        variables: List[str],
        correlations: Dict[Tuple[str, str], float],
        title: str = "Correlation Matrix",
        start_row: int = 1
    ) -> int:
        """
        Write APA-formatted correlation matrix.
        
        Args:
            variables: List of variable names.
            correlations: Dict mapping (var1, var2) to correlation.
            title: Table title.
            start_row: Starting row.
        
        Returns:
            Next available row.
        """
        row = start_row
        
        self.ws.cell(row=row, column=1, value="Table X")
        self.ws.cell(row=row, column=1).font = self.style.title_font
        row += 1
        
        self.ws.cell(row=row, column=1, value=title)
        self.ws.cell(row=row, column=1).font = Font(
            name='Times New Roman', size=12, italic=True
        )
        row += 2
        
        self.ws.cell(row=row, column=1, value="Variable").font = self.style.header_font
        for col, var in enumerate(variables, 2):
            cell = self.ws.cell(row=row, column=col, value=str(col - 1))
            cell.font = self.style.header_font
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        for i, var1 in enumerate(variables):
            self.ws.cell(row=row, column=1, value=f"{i + 1}. {var1}").font = self.style.body_font
            
            for j, var2 in enumerate(variables):
                col = j + 2
                if i == j:
                    val = "—"
                elif i > j:
                    val = ""
                else:
                    r = correlations.get((var1, var2), correlations.get((var2, var1), 0))
                    val = format_statistic(r, 2)
                
                cell = self.ws.cell(row=row, column=col, value=val)
                cell.font = self.style.body_font
                cell.alignment = Alignment(horizontal='center')
            row += 1
        
        row += 1
        note = "Note. Values above the diagonal represent Pearson correlations. *p < .05. **p < .01."
        self.ws.cell(row=row, column=1, value=note).font = self.style.note_font
        
        return row + 2

    def write_ttest_results(
        self,
        comparison_name: str,
        group1_name: str,
        group2_name: str,
        group1_stats: Dict[str, float],
        group2_stats: Dict[str, float],
        t_value: float,
        df: int,
        p_value: float,
        cohens_d: float,
        start_row: int = 1
    ) -> int:
        """
        Write APA-formatted t-test results.
        
        Args:
            comparison_name: Name of the comparison.
            group1_name: Name of first group.
            group2_name: Name of second group.
            group1_stats: {n, mean, sd} for group 1.
            group2_stats: {n, mean, sd} for group 2.
            t_value: t statistic.
            df: Degrees of freedom.
            p_value: P-value.
            cohens_d: Effect size.
            start_row: Starting row.
        
        Returns:
            Next available row.
        """
        row = start_row
        
        self.ws.cell(row=row, column=1, value="Table X")
        self.ws.cell(row=row, column=1).font = self.style.title_font
        row += 1
        
        self.ws.cell(row=row, column=1, value=f"Independent Samples t-Test: {comparison_name}")
        self.ws.cell(row=row, column=1).font = Font(
            name='Times New Roman', size=12, italic=True
        )
        row += 2
        
        headers = ["Group", "n", "M", "SD", "t", "df", "p", "d"]
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=row, column=col, value=header)
            cell.font = self.style.header_font
            cell.border = self.style.header_border
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        self.ws.cell(row=row, column=1, value=group1_name).font = self.style.body_font
        self.ws.cell(row=row, column=2, value=group1_stats.get('n', '')).font = self.style.body_font
        self.ws.cell(row=row, column=3, value=format_statistic(group1_stats.get('mean', 0))).font = self.style.body_font
        self.ws.cell(row=row, column=4, value=format_statistic(group1_stats.get('sd', 0))).font = self.style.body_font
        self.ws.cell(row=row, column=5, value=format_statistic(t_value)).font = self.style.body_font
        self.ws.cell(row=row, column=6, value=df).font = self.style.body_font
        self.ws.cell(row=row, column=7, value=format_p_value(p_value)).font = self.style.body_font
        self.ws.cell(row=row, column=8, value=format_statistic(cohens_d)).font = self.style.body_font
        row += 1
        
        self.ws.cell(row=row, column=1, value=group2_name).font = self.style.body_font
        self.ws.cell(row=row, column=2, value=group2_stats.get('n', '')).font = self.style.body_font
        self.ws.cell(row=row, column=3, value=format_statistic(group2_stats.get('mean', 0))).font = self.style.body_font
        self.ws.cell(row=row, column=4, value=format_statistic(group2_stats.get('sd', 0))).font = self.style.body_font
        row += 1
        
        row += 1
        effect_interp = interpret_cohens_d(cohens_d)
        note = f"Note. d = Cohen's d ({effect_interp.value} effect)."
        self.ws.cell(row=row, column=1, value=note).font = self.style.note_font
        
        return row + 2

    def write_reliability_results(
        self,
        scale_name: str,
        alpha: float,
        n_items: int,
        item_stats: Optional[List[Dict[str, Any]]] = None,
        start_row: int = 1
    ) -> int:
        """
        Write APA-formatted reliability analysis results.
        
        Args:
            scale_name: Name of the scale.
            alpha: Cronbach's alpha.
            n_items: Number of items.
            item_stats: Optional list of item-level statistics.
            start_row: Starting row.
        
        Returns:
            Next available row.
        """
        row = start_row
        
        self.ws.cell(row=row, column=1, value="Table X")
        self.ws.cell(row=row, column=1).font = self.style.title_font
        row += 1
        
        self.ws.cell(row=row, column=1, value=f"Reliability Analysis: {scale_name}")
        self.ws.cell(row=row, column=1).font = Font(
            name='Times New Roman', size=12, italic=True
        )
        row += 2
        
        interp = interpret_cronbach_alpha(alpha)
        self.ws.cell(row=row, column=1, value="Cronbach's α").font = self.style.header_font
        self.ws.cell(row=row, column=2, value=format_statistic(alpha, 3)).font = self.style.body_font
        self.ws.cell(row=row, column=3, value=f"({interp})").font = self.style.note_font
        row += 1
        
        self.ws.cell(row=row, column=1, value="Number of Items").font = self.style.header_font
        self.ws.cell(row=row, column=2, value=n_items).font = self.style.body_font
        row += 2
        
        if item_stats:
            headers = ["Item", "M", "SD", "Corrected Item-Total r", "α if Deleted"]
            for col, header in enumerate(headers, 1):
                cell = self.ws.cell(row=row, column=col, value=header)
                cell.font = self.style.header_font
                cell.border = self.style.header_border
            row += 1
            
            for item in item_stats:
                self.ws.cell(row=row, column=1, value=item.get('name', '')).font = self.style.body_font
                self.ws.cell(row=row, column=2, value=format_statistic(item.get('mean', 0))).font = self.style.body_font
                self.ws.cell(row=row, column=3, value=format_statistic(item.get('sd', 0))).font = self.style.body_font
                self.ws.cell(row=row, column=4, value=format_statistic(item.get('item_total_r', 0))).font = self.style.body_font
                self.ws.cell(row=row, column=5, value=format_statistic(item.get('alpha_if_deleted', 0), 3)).font = self.style.body_font
                row += 1
        
        return row + 2


def generate_apa_interpretation(
    stat_type: str,
    results: Dict[str, Any]
) -> str:
    """
    Generate APA-style interpretation text for statistical results.
    
    Args:
        stat_type: Type of statistic ('ttest', 'correlation', 'anova', etc.).
        results: Dictionary with statistical results.
    
    Returns:
        APA-formatted interpretation paragraph.
    """
    if stat_type == 'ttest':
        t = results.get('t', 0)
        df = results.get('df', 0)
        p = results.get('p', 1)
        d = results.get('d', 0)
        g1_name = results.get('group1_name', 'Group 1')
        g2_name = results.get('group2_name', 'Group 2')
        g1_m = results.get('group1_mean', 0)
        g2_m = results.get('group2_mean', 0)
        dv = results.get('dv_name', 'the dependent variable')
        
        sig = "was" if p < 0.05 else "was not"
        effect_desc = interpret_cohens_d(d).value
        
        return (
            f"An independent samples t-test was conducted to compare {dv} between "
            f"{g1_name} and {g2_name}. There {sig} a statistically significant difference "
            f"between groups, t({df}) = {t:.2f}, p {format_p_value(p)}, d = {d:.2f}. "
            f"{g1_name} (M = {g1_m:.2f}) scored {'higher' if g1_m > g2_m else 'lower'} than "
            f"{g2_name} (M = {g2_m:.2f}), representing a {effect_desc} effect size."
        )
    
    elif stat_type == 'correlation':
        r = results.get('r', 0)
        p = results.get('p', 1)
        n = results.get('n', 0)
        var1 = results.get('var1', 'Variable 1')
        var2 = results.get('var2', 'Variable 2')
        
        strength = interpret_correlation(r)
        sig = "was" if p < 0.05 else "was not"
        
        return (
            f"A Pearson correlation coefficient was computed to assess the relationship "
            f"between {var1} and {var2}. There {sig} a statistically significant "
            f"{strength} correlation between the two variables, r({n - 2}) = {r:.2f}, "
            f"p {format_p_value(p)}."
        )
    
    elif stat_type == 'reliability':
        alpha = results.get('alpha', 0)
        scale = results.get('scale_name', 'the scale')
        n_items = results.get('n_items', 0)
        
        interp = interpret_cronbach_alpha(alpha)
        
        return (
            f"Internal consistency for {scale} was assessed using Cronbach's alpha. "
            f"The {n_items}-item scale demonstrated {interp} internal consistency "
            f"(α = {alpha:.2f})."
        )
    
    return ""
