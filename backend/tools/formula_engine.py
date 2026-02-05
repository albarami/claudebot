"""
Deterministic Formula Engine.
Generates Excel formulas for each analysis task type.
All numeric outputs must be formulas referencing 00_RAW_DATA_LOCKED.
"""

from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass
from enum import Enum

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from models.task_schema import TaskType


RAW_DATA_SHEET = "00_RAW_DATA_LOCKED"


class FormulaType(str, Enum):
    """Types of formula outputs."""
    COUNT = "count"
    MEAN = "mean"
    MEDIAN = "median"
    STDEV = "stdev"
    VARIANCE = "variance"
    MIN = "min"
    MAX = "max"
    SUM = "sum"
    SKEWNESS = "skewness"
    KURTOSIS = "kurtosis"
    CORRELATION = "correlation"
    TTEST = "ttest"
    FREQUENCY = "frequency"
    PERCENTAGE = "percentage"
    MISSING = "missing"
    VALID_N = "valid_n"


@dataclass
class ColumnMapping:
    """Maps column names to Excel column letters."""
    name: str
    letter: str
    index: int
    data_start_row: int
    data_end_row: int

    @property
    def data_range(self) -> str:
        """Get the full data range for this column."""
        return f"'{RAW_DATA_SHEET}'!{self.letter}{self.data_start_row}:{self.letter}{self.data_end_row}"

    @property
    def absolute_range(self) -> str:
        """Get absolute reference range."""
        return f"'{RAW_DATA_SHEET}'!${self.letter}${self.data_start_row}:${self.letter}${self.data_end_row}"


@dataclass
class FormulaResult:
    """Result of formula generation."""
    cell: str
    formula: str
    label: str
    formula_type: FormulaType


class FormulaEngine:
    """
    Deterministic formula generator for survey analysis.
    All formulas reference the raw data sheet.
    """

    def __init__(
        self,
        columns: List[str],
        n_rows: int,
        header_row: int = 1,
        data_start_row: int = 2
    ):
        """
        Initialize the formula engine.

        Args:
            columns: List of column names in order.
            n_rows: Total number of data rows (excluding header).
            header_row: Row number containing headers.
            data_start_row: First row of data.
        """
        self.columns = columns
        self.n_rows = n_rows
        self.header_row = header_row
        self.data_start_row = data_start_row
        self.data_end_row = data_start_row + n_rows - 1

        self.column_map: Dict[str, ColumnMapping] = {}
        for idx, col_name in enumerate(columns):
            letter = get_column_letter(idx + 1)
            self.column_map[col_name] = ColumnMapping(
                name=col_name,
                letter=letter,
                index=idx + 1,
                data_start_row=data_start_row,
                data_end_row=self.data_end_row
            )

    def get_column(self, name: str) -> Optional[ColumnMapping]:
        """Get column mapping by name."""
        return self.column_map.get(name)

    def get_column_range(self, col_name: str) -> str:
        """Get the data range for a column."""
        col = self.get_column(col_name)
        if col:
            return col.data_range
        raise ValueError(f"Column '{col_name}' not found")

    def generate_descriptive_formulas(
        self,
        col_name: str,
        output_col: int,
        output_start_row: int
    ) -> List[FormulaResult]:
        """
        Generate descriptive statistics formulas for a column.

        Args:
            col_name: Column to analyze.
            output_col: Output column number (1-indexed).
            output_start_row: Starting row for output.

        Returns:
            List of formula results.
        """
        col = self.get_column(col_name)
        if not col:
            raise ValueError(f"Column '{col_name}' not found")

        data_range = col.data_range
        out_letter = get_column_letter(output_col)
        results = []

        formulas = [
            (FormulaType.COUNT, "N", f"=COUNT({data_range})"),
            (FormulaType.VALID_N, "Valid N", f"=COUNT({data_range})"),
            (FormulaType.MISSING, "Missing", f"=COUNTBLANK({data_range})"),
            (FormulaType.MEAN, "Mean", f"=AVERAGE({data_range})"),
            (FormulaType.MEDIAN, "Median", f"=MEDIAN({data_range})"),
            (FormulaType.STDEV, "Std. Deviation", f"=STDEV.S({data_range})"),
            (FormulaType.VARIANCE, "Variance", f"=VAR.S({data_range})"),
            (FormulaType.MIN, "Minimum", f"=MIN({data_range})"),
            (FormulaType.MAX, "Maximum", f"=MAX({data_range})"),
            (FormulaType.SKEWNESS, "Skewness", f"=SKEW({data_range})"),
            (FormulaType.KURTOSIS, "Kurtosis", f"=KURT({data_range})"),
        ]

        for row_offset, (ftype, label, formula) in enumerate(formulas):
            row = output_start_row + row_offset
            results.append(FormulaResult(
                cell=f"{out_letter}{row}",
                formula=formula,
                label=label,
                formula_type=ftype
            ))

        return results

    def generate_frequency_formulas(
        self,
        col_name: str,
        unique_values: List[Any],
        output_col: int,
        output_start_row: int
    ) -> List[FormulaResult]:
        """
        Generate frequency table formulas.

        Args:
            col_name: Column to analyze.
            unique_values: List of unique values in the column.
            output_col: Output column for counts (1-indexed).
            output_start_row: Starting row.

        Returns:
            List of formula results.
        """
        col = self.get_column(col_name)
        if not col:
            raise ValueError(f"Column '{col_name}' not found")

        data_range = col.data_range
        out_letter = get_column_letter(output_col)
        pct_letter = get_column_letter(output_col + 1)
        results = []

        total_formula = f"=COUNTA({data_range})"
        total_row = output_start_row + len(unique_values)

        for row_offset, value in enumerate(unique_values):
            row = output_start_row + row_offset

            if isinstance(value, str):
                count_formula = f'=COUNTIF({data_range},"{value}")'
            else:
                count_formula = f"=COUNTIF({data_range},{value})"

            results.append(FormulaResult(
                cell=f"{out_letter}{row}",
                formula=count_formula,
                label=f"Count: {value}",
                formula_type=FormulaType.FREQUENCY
            ))

            pct_formula = f"={out_letter}{row}/{out_letter}{total_row}*100"
            results.append(FormulaResult(
                cell=f"{pct_letter}{row}",
                formula=pct_formula,
                label=f"Percent: {value}",
                formula_type=FormulaType.PERCENTAGE
            ))

        results.append(FormulaResult(
            cell=f"{out_letter}{total_row}",
            formula=total_formula,
            label="Total",
            formula_type=FormulaType.COUNT
        ))

        return results

    def generate_correlation_formula(
        self,
        col1_name: str,
        col2_name: str
    ) -> str:
        """
        Generate correlation formula between two columns.

        Args:
            col1_name: First column name.
            col2_name: Second column name.

        Returns:
            Excel CORREL formula string.
        """
        col1 = self.get_column(col1_name)
        col2 = self.get_column(col2_name)
        if not col1 or not col2:
            raise ValueError("Column not found")

        return f"=CORREL({col1.data_range},{col2.data_range})"

    def generate_ttest_formula(
        self,
        col1_name: str,
        col2_name: str,
        tails: int = 2,
        test_type: int = 2
    ) -> str:
        """
        Generate T.TEST formula.

        Args:
            col1_name: First column (group 1).
            col2_name: Second column (group 2).
            tails: 1 for one-tailed, 2 for two-tailed.
            test_type: 1=paired, 2=equal variance, 3=unequal variance.

        Returns:
            Excel T.TEST formula string.
        """
        col1 = self.get_column(col1_name)
        col2 = self.get_column(col2_name)
        if not col1 or not col2:
            raise ValueError("Column not found")

        return f"=T.TEST({col1.data_range},{col2.data_range},{tails},{test_type})"

    def generate_grouped_mean_formula(
        self,
        value_col: str,
        group_col: str,
        group_value: Any
    ) -> str:
        """
        Generate AVERAGEIF formula for grouped mean.

        Args:
            value_col: Column containing values to average.
            group_col: Column containing group labels.
            group_value: Value identifying the group.

        Returns:
            Excel AVERAGEIF formula.
        """
        val_col = self.get_column(value_col)
        grp_col = self.get_column(group_col)
        if not val_col or not grp_col:
            raise ValueError("Column not found")

        if isinstance(group_value, str):
            criteria = f'"{group_value}"'
        else:
            criteria = str(group_value)

        return f"=AVERAGEIF({grp_col.data_range},{criteria},{val_col.data_range})"

    def generate_grouped_count_formula(
        self,
        group_col: str,
        group_value: Any
    ) -> str:
        """
        Generate COUNTIF formula for group count.

        Args:
            group_col: Column containing group labels.
            group_value: Value identifying the group.

        Returns:
            Excel COUNTIF formula.
        """
        grp_col = self.get_column(group_col)
        if not grp_col:
            raise ValueError(f"Column '{group_col}' not found")

        if isinstance(group_value, str):
            criteria = f'"{group_value}"'
        else:
            criteria = str(group_value)

        return f"=COUNTIF({grp_col.data_range},{criteria})"

    def generate_grouped_stdev_formula(
        self,
        value_col: str,
        group_col: str,
        group_value: Any
    ) -> str:
        """
        Generate array formula for grouped standard deviation.
        Note: Requires Ctrl+Shift+Enter or dynamic array support.

        Args:
            value_col: Column containing values.
            group_col: Column containing group labels.
            group_value: Value identifying the group.

        Returns:
            Excel array formula for grouped STDEV.
        """
        val_col = self.get_column(value_col)
        grp_col = self.get_column(group_col)
        if not val_col or not grp_col:
            raise ValueError("Column not found")

        if isinstance(group_value, str):
            criteria = f'"{group_value}"'
        else:
            criteria = str(group_value)

        return f"=STDEV.S(IF({grp_col.data_range}={criteria},{val_col.data_range}))"

    def generate_missing_analysis_formulas(
        self,
        col_names: List[str],
        output_start_row: int
    ) -> List[FormulaResult]:
        """
        Generate missing data analysis formulas.

        Args:
            col_names: Columns to analyze.
            output_start_row: Starting row for output.

        Returns:
            List of formula results for missing analysis.
        """
        results = []

        for row_offset, col_name in enumerate(col_names):
            col = self.get_column(col_name)
            if not col:
                continue

            row = output_start_row + row_offset
            data_range = col.data_range

            results.append(FormulaResult(
                cell=f"B{row}",
                formula=f"=COUNT({data_range})",
                label=f"Valid N: {col_name}",
                formula_type=FormulaType.VALID_N
            ))

            results.append(FormulaResult(
                cell=f"C{row}",
                formula=f"=COUNTBLANK({data_range})",
                label=f"Missing: {col_name}",
                formula_type=FormulaType.MISSING
            ))

            total_cells = self.n_rows
            results.append(FormulaResult(
                cell=f"D{row}",
                formula=f"=C{row}/{total_cells}*100",
                label=f"Missing %: {col_name}",
                formula_type=FormulaType.PERCENTAGE
            ))

        return results

    def generate_cronbach_alpha_formulas(
        self,
        item_cols: List[str],
        output_row: int
    ) -> List[FormulaResult]:
        """
        Generate formulas for Cronbach's alpha calculation.
        Uses UDF if available, otherwise provides component formulas.

        Args:
            item_cols: List of scale item column names.
            output_row: Row for output.

        Returns:
            List of formula results.
        """
        if not item_cols or len(item_cols) < 2:
            raise ValueError("At least 2 items required for reliability")

        results = []
        k = len(item_cols)

        col_ranges = []
        for col_name in item_cols:
            col = self.get_column(col_name)
            if col:
                col_ranges.append(col.data_range)

        first_col = self.get_column(item_cols[0])
        last_col = self.get_column(item_cols[-1])
        if first_col and last_col:
            full_range = (
                f"'{RAW_DATA_SHEET}'!"
                f"{first_col.letter}{self.data_start_row}:"
                f"{last_col.letter}{self.data_end_row}"
            )

            results.append(FormulaResult(
                cell=f"B{output_row}",
                formula=f"=CRONBACH_ALPHA({full_range})",
                label="Cronbach's Alpha (UDF)",
                formula_type=FormulaType.COUNT
            ))

        results.append(FormulaResult(
            cell=f"C{output_row}",
            formula=f"={k}",
            label="Number of Items",
            formula_type=FormulaType.COUNT
        ))

        return results

    def generate_effect_size_formulas(
        self,
        mean1_cell: str,
        mean2_cell: str,
        sd1_cell: str,
        sd2_cell: str,
        n1_cell: str,
        n2_cell: str,
        output_row: int
    ) -> List[FormulaResult]:
        """
        Generate Cohen's d effect size formula.

        Args:
            mean1_cell: Cell reference for group 1 mean.
            mean2_cell: Cell reference for group 2 mean.
            sd1_cell: Cell reference for group 1 SD.
            sd2_cell: Cell reference for group 2 SD.
            n1_cell: Cell reference for group 1 N.
            n2_cell: Cell reference for group 2 N.
            output_row: Row for output.

        Returns:
            List of formula results.
        """
        results = []

        pooled_sd_formula = (
            f"=SQRT((({n1_cell}-1)*{sd1_cell}^2+({n2_cell}-1)*{sd2_cell}^2)"
            f"/({n1_cell}+{n2_cell}-2))"
        )
        results.append(FormulaResult(
            cell=f"E{output_row}",
            formula=pooled_sd_formula,
            label="Pooled SD",
            formula_type=FormulaType.STDEV
        ))

        cohens_d_formula = f"=({mean1_cell}-{mean2_cell})/E{output_row}"
        results.append(FormulaResult(
            cell=f"F{output_row}",
            formula=cohens_d_formula,
            label="Cohen's d",
            formula_type=FormulaType.COUNT
        ))

        udf_formula = f"=COHENS_D({mean1_cell},{sd1_cell},{n1_cell},{mean2_cell},{sd2_cell},{n2_cell})"
        results.append(FormulaResult(
            cell=f"G{output_row}",
            formula=udf_formula,
            label="Cohen's d (UDF)",
            formula_type=FormulaType.COUNT
        ))

        return results


def write_formulas_to_sheet(
    ws: Worksheet,
    formulas: List[FormulaResult],
    include_labels: bool = True,
    label_col: int = 1
) -> None:
    """
    Write formula results to a worksheet.

    Args:
        ws: Target worksheet.
        formulas: List of formula results.
        include_labels: Whether to write labels.
        label_col: Column for labels (1-indexed).
    """
    for fr in formulas:
        col_letter = fr.cell[0] if fr.cell[0].isalpha() else fr.cell[:2]
        row = int(''.join(c for c in fr.cell if c.isdigit()))

        ws[fr.cell] = fr.formula

        if include_labels:
            label_letter = get_column_letter(label_col)
            ws[f"{label_letter}{row}"] = fr.label


def create_formula_engine(
    columns: List[str],
    n_rows: int
) -> FormulaEngine:
    """
    Factory function to create a formula engine.

    Args:
        columns: List of column names.
        n_rows: Number of data rows.

    Returns:
        Configured FormulaEngine instance.
    """
    return FormulaEngine(columns=columns, n_rows=n_rows)
