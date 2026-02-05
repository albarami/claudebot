"""
Deterministic Verification Module.
Computes ground-truth statistics using Python and compares to Excel output.
Verification must pass for QC approval.
"""

from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
import math

import numpy as np
import pandas as pd
from scipy import stats
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter


class VerificationStatus(str, Enum):
    """Verification result status."""
    PASS = "PASS"
    FAIL = "FAIL"
    SKIP = "SKIP"
    ERROR = "ERROR"


@dataclass
class VerificationCheck:
    """Single verification check result."""
    check_name: str
    expected_value: float
    actual_value: Optional[float]
    tolerance: float
    status: VerificationStatus
    cell_reference: str
    details: str = ""

    @property
    def difference(self) -> Optional[float]:
        """Calculate absolute difference."""
        if self.actual_value is None:
            return None
        return abs(self.expected_value - self.actual_value)

    @property
    def within_tolerance(self) -> bool:
        """Check if difference is within tolerance."""
        diff = self.difference
        if diff is None:
            return False
        return diff <= self.tolerance


@dataclass
class VerificationResult:
    """Complete verification result for a task or sheet."""
    task_id: str
    sheet_name: str
    status: VerificationStatus
    checks: List[VerificationCheck] = field(default_factory=list)
    formula_coverage: float = 0.0
    errors: List[str] = field(default_factory=list)

    @property
    def passed_checks(self) -> int:
        """Count of passed checks."""
        return sum(1 for c in self.checks if c.status == VerificationStatus.PASS)

    @property
    def failed_checks(self) -> int:
        """Count of failed checks."""
        return sum(1 for c in self.checks if c.status == VerificationStatus.FAIL)

    @property
    def total_checks(self) -> int:
        """Total number of checks."""
        return len(self.checks)

    @property
    def pass_rate(self) -> float:
        """Percentage of checks passed."""
        if self.total_checks == 0:
            return 0.0
        return (self.passed_checks / self.total_checks) * 100


DEFAULT_TOLERANCE = 1e-6
PERCENTAGE_TOLERANCE = 0.01
STATISTICAL_TOLERANCE = 1e-4


class StatisticalVerifier:
    """
    Computes ground-truth statistics for verification.
    Uses scipy and numpy for accurate calculations.
    """

    def __init__(self, data: pd.DataFrame):
        """
        Initialize with source data.

        Args:
            data: DataFrame with raw survey data.
        """
        self.data = data

    def compute_descriptives(self, column: str) -> Dict[str, float]:
        """
        Compute descriptive statistics for a column.

        Args:
            column: Column name.

        Returns:
            Dict with count, mean, std, min, max, skew, kurtosis.
        """
        series = self.data[column].dropna()

        if len(series) == 0:
            return {
                "count": 0,
                "mean": float('nan'),
                "std": float('nan'),
                "min": float('nan'),
                "max": float('nan'),
                "skewness": float('nan'),
                "kurtosis": float('nan'),
                "median": float('nan'),
                "variance": float('nan'),
                "missing": len(self.data[column]) - len(series)
            }

        return {
            "count": len(series),
            "mean": series.mean(),
            "std": series.std(ddof=1),
            "min": series.min(),
            "max": series.max(),
            "skewness": stats.skew(series, bias=False) if len(series) >= 3 else float('nan'),
            "kurtosis": stats.kurtosis(series, bias=False) if len(series) >= 4 else float('nan'),
            "median": series.median(),
            "variance": series.var(ddof=1),
            "missing": len(self.data[column]) - len(series)
        }

    def compute_correlation(self, col1: str, col2: str) -> float:
        """
        Compute Pearson correlation.

        Args:
            col1: First column.
            col2: Second column.

        Returns:
            Correlation coefficient.
        """
        valid = self.data[[col1, col2]].dropna()
        if len(valid) < 3:
            return float('nan')
        return valid[col1].corr(valid[col2])

    def compute_ttest(
        self,
        col1: str,
        col2: str,
        paired: bool = False
    ) -> Tuple[float, float]:
        """
        Compute t-test.

        Args:
            col1: First group/column.
            col2: Second group/column.
            paired: Whether test is paired.

        Returns:
            Tuple of (t-statistic, p-value).
        """
        data1 = self.data[col1].dropna()
        data2 = self.data[col2].dropna()

        if len(data1) < 2 or len(data2) < 2:
            return (float('nan'), float('nan'))

        if paired:
            result = stats.ttest_rel(data1, data2)
        else:
            result = stats.ttest_ind(data1, data2, equal_var=True)

        return (result.statistic, result.pvalue)

    def compute_frequency(self, column: str) -> Dict[Any, int]:
        """
        Compute frequency counts.

        Args:
            column: Column name.

        Returns:
            Dict mapping values to counts.
        """
        return self.data[column].value_counts().to_dict()

    def compute_cronbach_alpha(self, columns: List[str]) -> float:
        """
        Compute Cronbach's alpha.

        Args:
            columns: List of scale item columns.

        Returns:
            Alpha coefficient.
        """
        subset = self.data[columns].dropna()
        if len(subset) < 2 or len(columns) < 2:
            return float('nan')

        k = len(columns)
        item_variances = subset.var(ddof=1)
        total_variance = subset.sum(axis=1).var(ddof=1)

        if total_variance == 0:
            return 0.0

        alpha = (k / (k - 1)) * (1 - item_variances.sum() / total_variance)
        return alpha

    def compute_shapiro_wilk(self, column: str) -> Tuple[float, float]:
        """
        Compute Shapiro-Wilk test.

        Args:
            column: Column name.

        Returns:
            Tuple of (W statistic, p-value).
        """
        series = self.data[column].dropna()
        if len(series) < 3 or len(series) > 5000:
            return (float('nan'), float('nan'))

        result = stats.shapiro(series)
        return (result.statistic, result.pvalue)

    def compute_levene(self, *groups: pd.Series) -> Tuple[float, float]:
        """
        Compute Levene's test.

        Args:
            groups: Variable number of group series.

        Returns:
            Tuple of (F statistic, p-value).
        """
        valid_groups = [g.dropna() for g in groups if len(g.dropna()) >= 2]
        if len(valid_groups) < 2:
            return (float('nan'), float('nan'))

        result = stats.levene(*valid_groups)
        return (result.statistic, result.pvalue)

    def compute_cohens_d(
        self,
        group1: pd.Series,
        group2: pd.Series
    ) -> float:
        """
        Compute Cohen's d effect size.

        Args:
            group1: First group data.
            group2: Second group data.

        Returns:
            Cohen's d value.
        """
        g1 = group1.dropna()
        g2 = group2.dropna()

        if len(g1) < 2 or len(g2) < 2:
            return float('nan')

        n1, n2 = len(g1), len(g2)
        var1, var2 = g1.var(ddof=1), g2.var(ddof=1)
        pooled_std = math.sqrt(((n1 - 1) * var1 + (n2 - 1) * var2) / (n1 + n2 - 2))

        if pooled_std == 0:
            return 0.0

        return (g1.mean() - g2.mean()) / pooled_std


class ExcelVerifier:
    """
    Verifies Excel workbook against Python ground truth.
    """

    def __init__(
        self,
        workbook_path: Path,
        raw_data: pd.DataFrame
    ):
        """
        Initialize verifier.

        Args:
            workbook_path: Path to Excel workbook.
            raw_data: Original data for ground truth.
        """
        self.workbook_path = workbook_path
        self.raw_data = raw_data
        self.stats = StatisticalVerifier(raw_data)

        if REQUIRE_EXCEL_RECALC:
            recalculate_workbook(workbook_path)

        keep_vba = workbook_path.suffix.lower() == '.xlsm'
        self.workbook = load_workbook(
            workbook_path,
            data_only=True,
            keep_vba=keep_vba
        )

    def close(self) -> None:
        """Close workbook."""
        if self.workbook:
            self.workbook.close()

    def verify_formula_coverage(
        self,
        sheet_name: str,
        data_region: Tuple[int, int, int, int]
    ) -> float:
        """
        Check percentage of numeric cells that contain formulas.

        Args:
            sheet_name: Sheet to check.
            data_region: (start_row, start_col, end_row, end_col).

        Returns:
            Percentage of formula-based numeric cells.
        """
        formula_wb = load_workbook(self.workbook_path, data_only=False)
        if sheet_name not in formula_wb.sheetnames:
            formula_wb.close()
            return 0.0

        ws = formula_wb[sheet_name]
        start_row, start_col, end_row, end_col = data_region

        total_numeric = 0
        formula_count = 0

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)) or (
                    isinstance(cell.value, str) and cell.value.startswith('=')
                ):
                    total_numeric += 1
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1

        formula_wb.close()

        if total_numeric == 0:
            return 100.0
        return (formula_count / total_numeric) * 100

    def verify_descriptives(
        self,
        sheet_name: str,
        column_name: str,
        cell_map: Dict[str, str]
    ) -> List[VerificationCheck]:
        """
        Verify descriptive statistics.

        Args:
            sheet_name: Sheet containing results.
            column_name: Column that was analyzed.
            cell_map: Map of stat names to cell references.

        Returns:
            List of verification checks.
        """
        if sheet_name not in self.workbook.sheetnames:
            return [VerificationCheck(
                check_name=f"Sheet existence: {sheet_name}",
                expected_value=1,
                actual_value=0,
                tolerance=0,
                status=VerificationStatus.FAIL,
                cell_reference="N/A",
                details="Sheet not found"
            )]

        ws = self.workbook[sheet_name]
        expected = self.stats.compute_descriptives(column_name)
        checks = []

        stat_tolerances = {
            "count": 0,
            "mean": STATISTICAL_TOLERANCE,
            "std": STATISTICAL_TOLERANCE,
            "min": STATISTICAL_TOLERANCE,
            "max": STATISTICAL_TOLERANCE,
            "skewness": 0.001,
            "kurtosis": 0.001,
            "median": STATISTICAL_TOLERANCE,
            "variance": STATISTICAL_TOLERANCE,
            "missing": 0
        }

        for stat_name, cell_ref in cell_map.items():
            if stat_name not in expected:
                continue

            exp_val = expected[stat_name]
            actual_val = ws[cell_ref].value
            tolerance = stat_tolerances.get(stat_name, DEFAULT_TOLERANCE)

            if actual_val is None:
                status = VerificationStatus.FAIL
                details = "Cell is empty"
            elif math.isnan(exp_val):
                status = VerificationStatus.SKIP
                details = "Expected value is NaN"
            elif isinstance(actual_val, (int, float)):
                diff = abs(exp_val - actual_val)
                if diff <= tolerance:
                    status = VerificationStatus.PASS
                    details = f"Difference: {diff:.6f}"
                else:
                    status = VerificationStatus.FAIL
                    details = f"Difference {diff:.6f} exceeds tolerance {tolerance}"
            else:
                status = VerificationStatus.FAIL
                details = f"Non-numeric value: {actual_val}"

            checks.append(VerificationCheck(
                check_name=f"{column_name}.{stat_name}",
                expected_value=exp_val,
                actual_value=float(actual_val) if isinstance(actual_val, (int, float)) else None,
                tolerance=tolerance,
                status=status,
                cell_reference=cell_ref,
                details=details
            ))

        return checks

    def verify_correlation_matrix(
        self,
        sheet_name: str,
        columns: List[str],
        start_row: int,
        start_col: int
    ) -> List[VerificationCheck]:
        """
        Verify correlation matrix values.

        Args:
            sheet_name: Sheet containing matrix.
            columns: List of columns in matrix.
            start_row: Starting row of matrix values.
            start_col: Starting column of matrix values.

        Returns:
            List of verification checks.
        """
        if sheet_name not in self.workbook.sheetnames:
            return [VerificationCheck(
                check_name="Sheet existence",
                expected_value=1,
                actual_value=0,
                tolerance=0,
                status=VerificationStatus.FAIL,
                cell_reference="N/A",
                details=f"Sheet {sheet_name} not found"
            )]

        ws = self.workbook[sheet_name]
        checks = []

        for i, col1 in enumerate(columns):
            for j, col2 in enumerate(columns):
                if j <= i:
                    continue

                exp_corr = self.stats.compute_correlation(col1, col2)
                cell_row = start_row + i
                cell_col = start_col + j
                cell_ref = f"{get_column_letter(cell_col)}{cell_row}"
                actual_val = ws.cell(row=cell_row, column=cell_col).value

                if math.isnan(exp_corr):
                    status = VerificationStatus.SKIP
                    details = "Cannot compute correlation"
                elif actual_val is None:
                    status = VerificationStatus.FAIL
                    details = "Cell is empty"
                elif isinstance(actual_val, (int, float)):
                    diff = abs(exp_corr - actual_val)
                    if diff <= STATISTICAL_TOLERANCE:
                        status = VerificationStatus.PASS
                        details = f"Diff: {diff:.6f}"
                    else:
                        status = VerificationStatus.FAIL
                        details = f"Diff {diff:.6f} > {STATISTICAL_TOLERANCE}"
                else:
                    status = VerificationStatus.FAIL
                    details = "Non-numeric value"

                checks.append(VerificationCheck(
                    check_name=f"r({col1},{col2})",
                    expected_value=exp_corr,
                    actual_value=float(actual_val) if isinstance(actual_val, (int, float)) else None,
                    tolerance=STATISTICAL_TOLERANCE,
                    status=status,
                    cell_reference=cell_ref,
                    details=details
                ))

        return checks


def verify_task_output(
    workbook_path: Path,
    raw_data: pd.DataFrame,
    task_id: str,
    task_type: str,
    sheet_name: str,
    verification_config: Dict[str, Any]
) -> VerificationResult:
    """
    Verify a single task's output.

    Args:
        workbook_path: Path to Excel workbook.
        raw_data: Original DataFrame.
        task_id: Task identifier.
        task_type: Type of task (from TaskType enum).
        sheet_name: Output sheet name.
        verification_config: Task-specific verification parameters.

    Returns:
        VerificationResult with all checks.
    """
    verifier = ExcelVerifier(workbook_path, raw_data)

    try:
        checks = []

        if task_type == "descriptive_stats":
            columns = verification_config.get("columns", [])
            cell_maps = verification_config.get("cell_maps", {})
            for col in columns:
                if col in cell_maps:
                    checks.extend(verifier.verify_descriptives(
                        sheet_name, col, cell_maps[col]
                    ))

        elif task_type == "correlation_matrix":
            columns = verification_config.get("columns", [])
            start_row = verification_config.get("start_row", 2)
            start_col = verification_config.get("start_col", 2)
            checks.extend(verifier.verify_correlation_matrix(
                sheet_name, columns, start_row, start_col
            ))

        data_region = verification_config.get("data_region", (2, 2, 50, 10))
        formula_coverage = verifier.verify_formula_coverage(sheet_name, data_region)

        failed = sum(1 for c in checks if c.status == VerificationStatus.FAIL)
        if failed > 0 or formula_coverage < 90:
            status = VerificationStatus.FAIL
        else:
            status = VerificationStatus.PASS

        result = VerificationResult(
            task_id=task_id,
            sheet_name=sheet_name,
            status=status,
            checks=checks,
            formula_coverage=formula_coverage
        )

    finally:
        verifier.close()

    return result


def generate_verification_report(results: List[VerificationResult]) -> str:
    """
    Generate human-readable verification report.

    Args:
        results: List of verification results.

    Returns:
        Formatted report string.
    """
    lines = [
        "=" * 60,
        "VERIFICATION REPORT",
        "=" * 60,
        ""
    ]

    total_pass = sum(r.passed_checks for r in results)
    total_fail = sum(r.failed_checks for r in results)
    total_checks = sum(r.total_checks for r in results)

    lines.append(f"Total Checks: {total_checks}")
    lines.append(f"Passed: {total_pass}")
    lines.append(f"Failed: {total_fail}")
    lines.append(f"Pass Rate: {(total_pass/total_checks*100) if total_checks > 0 else 0:.1f}%")
    lines.append("")

    for result in results:
        status_icon = "✓" if result.status == VerificationStatus.PASS else "✗"
        lines.append(f"{status_icon} Task {result.task_id}: {result.sheet_name}")
        lines.append(f"   Formula Coverage: {result.formula_coverage:.1f}%")
        lines.append(f"   Checks: {result.passed_checks}/{result.total_checks} passed")

        if result.failed_checks > 0:
            lines.append("   Failed checks:")
            for check in result.checks:
                if check.status == VerificationStatus.FAIL:
                    lines.append(
                        f"     - {check.check_name}: "
                        f"expected {check.expected_value:.4f}, "
                        f"got {check.actual_value}"
                    )
        lines.append("")

    lines.append("=" * 60)
    overall = "PASS" if total_fail == 0 else "FAIL"
    lines.append(f"OVERALL: {overall}")
    lines.append("=" * 60)

    return "\n".join(lines)
from config import REQUIRE_EXCEL_RECALC
from tools.excel_com import recalculate_workbook
