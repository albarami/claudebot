"""
Qualitative Analysis Tools.
Automated codebook creation, coding, and inter-rater reliability.
"""

from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, field
from collections import Counter
import re
import hashlib

import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter


@dataclass
class Code:
    """A single code in the codebook."""
    id: str
    name: str
    definition: str
    examples: List[str] = field(default_factory=list)
    parent_code: Optional[str] = None
    frequency: int = 0


@dataclass
class Theme:
    """A theme grouping multiple codes."""
    id: str
    name: str
    description: str
    codes: List[str] = field(default_factory=list)


@dataclass
class Codebook:
    """Complete codebook for qualitative analysis."""
    name: str
    codes: Dict[str, Code] = field(default_factory=dict)
    themes: Dict[str, Theme] = field(default_factory=dict)
    
    def add_code(self, code: Code) -> None:
        """Add a code to the codebook."""
        self.codes[code.id] = code
    
    def add_theme(self, theme: Theme) -> None:
        """Add a theme to the codebook."""
        self.themes[theme.id] = theme
    
    def get_codes_by_theme(self, theme_id: str) -> List[Code]:
        """Get all codes belonging to a theme."""
        theme = self.themes.get(theme_id)
        if not theme:
            return []
        return [self.codes[cid] for cid in theme.codes if cid in self.codes]


@dataclass
class CodingResult:
    """Result of coding a single response."""
    response_id: str
    response_text: str
    assigned_codes: List[str]
    coder_id: str
    confidence: float = 1.0


class AutomatedCoder:
    """
    Automated qualitative coder using keyword and pattern matching.
    For PhD-level analysis, this should be supplemented with LLM coding.
    """
    
    def __init__(self, codebook: Codebook):
        self.codebook = codebook
        self.code_patterns: Dict[str, List[re.Pattern]] = {}
        self._compile_patterns()
    
    def _compile_patterns(self) -> None:
        """Compile regex patterns for each code based on name and examples."""
        for code_id, code in self.codebook.codes.items():
            patterns = []
            
            name_words = code.name.lower().split()
            for word in name_words:
                if len(word) > 3:
                    patterns.append(re.compile(rf'\b{re.escape(word)}\w*\b', re.IGNORECASE))
            
            for example in code.examples:
                example_clean = re.escape(example.lower())
                patterns.append(re.compile(rf'\b{example_clean}\b', re.IGNORECASE))
            
            self.code_patterns[code_id] = patterns
    
    def code_response(
        self,
        response_id: str,
        response_text: str,
        coder_id: str = "auto_coder_1"
    ) -> CodingResult:
        """
        Code a single response.
        
        Args:
            response_id: Unique identifier for the response.
            response_text: Text to code.
            coder_id: Identifier for this coder.
        
        Returns:
            CodingResult with assigned codes.
        """
        assigned_codes = []
        
        for code_id, patterns in self.code_patterns.items():
            match_count = 0
            for pattern in patterns:
                if pattern.search(response_text):
                    match_count += 1
            
            if match_count > 0:
                assigned_codes.append(code_id)
        
        confidence = min(1.0, len(assigned_codes) / max(len(self.codebook.codes) * 0.1, 1))
        
        return CodingResult(
            response_id=response_id,
            response_text=response_text,
            assigned_codes=assigned_codes,
            coder_id=coder_id,
            confidence=confidence
        )
    
    def code_dataframe_column(
        self,
        df: pd.DataFrame,
        column: str,
        coder_id: str = "auto_coder_1"
    ) -> List[CodingResult]:
        """
        Code all responses in a DataFrame column.
        
        Args:
            df: DataFrame with responses.
            column: Column name containing text responses.
            coder_id: Coder identifier.
        
        Returns:
            List of CodingResults.
        """
        results = []
        for idx, row in df.iterrows():
            text = str(row[column]) if pd.notna(row[column]) else ""
            if text and text.lower() != 'nan':
                result = self.code_response(
                    response_id=str(idx),
                    response_text=text,
                    coder_id=coder_id
                )
                results.append(result)
        return results


def calculate_cohens_kappa(
    coder1_results: List[CodingResult],
    coder2_results: List[CodingResult],
    codes: List[str]
) -> float:
    """
    Calculate Cohen's kappa for inter-rater reliability.
    
    Args:
        coder1_results: Results from first coder.
        coder2_results: Results from second coder.
        codes: List of all possible codes.
    
    Returns:
        Cohen's kappa coefficient.
    """
    coder1_map = {r.response_id: set(r.assigned_codes) for r in coder1_results}
    coder2_map = {r.response_id: set(r.assigned_codes) for r in coder2_results}
    
    common_ids = set(coder1_map.keys()) & set(coder2_map.keys())
    if not common_ids:
        return 0.0
    
    agreements = 0
    total_decisions = 0
    
    for response_id in common_ids:
        codes1 = coder1_map[response_id]
        codes2 = coder2_map[response_id]
        
        for code in codes:
            in_coder1 = code in codes1
            in_coder2 = code in codes2
            if in_coder1 == in_coder2:
                agreements += 1
            total_decisions += 1
    
    if total_decisions == 0:
        return 0.0
    
    p_o = agreements / total_decisions
    
    coder1_positive = sum(1 for rid in common_ids for c in codes if c in coder1_map[rid])
    coder2_positive = sum(1 for rid in common_ids for c in codes if c in coder2_map[rid])
    total_possible = len(common_ids) * len(codes)
    
    p1_pos = coder1_positive / total_possible if total_possible > 0 else 0
    p2_pos = coder2_positive / total_possible if total_possible > 0 else 0
    
    p_e = (p1_pos * p2_pos) + ((1 - p1_pos) * (1 - p2_pos))
    
    if p_e == 1:
        return 1.0
    
    kappa = (p_o - p_e) / (1 - p_e)
    return kappa


def interpret_kappa(kappa: float) -> str:
    """
    Interpret Cohen's kappa value.
    
    Args:
        kappa: Kappa coefficient.
    
    Returns:
        Interpretation string.
    """
    if kappa < 0:
        return "poor (less than chance)"
    elif kappa < 0.20:
        return "slight"
    elif kappa < 0.40:
        return "fair"
    elif kappa < 0.60:
        return "moderate"
    elif kappa < 0.80:
        return "substantial"
    else:
        return "almost perfect"


def generate_frequency_table(
    coding_results: List[CodingResult],
    codebook: Codebook
) -> pd.DataFrame:
    """
    Generate frequency table for codes.
    
    Args:
        coding_results: List of coding results.
        codebook: Codebook with code definitions.
    
    Returns:
        DataFrame with code frequencies.
    """
    code_counts = Counter()
    for result in coding_results:
        for code_id in result.assigned_codes:
            code_counts[code_id] += 1
    
    total_responses = len(coding_results)
    
    rows = []
    for code_id, code in codebook.codes.items():
        count = code_counts.get(code_id, 0)
        percentage = (count / total_responses * 100) if total_responses > 0 else 0
        rows.append({
            'Code ID': code_id,
            'Code Name': code.name,
            'Definition': code.definition[:100],
            'Frequency': count,
            'Percentage': round(percentage, 1)
        })
    
    df = pd.DataFrame(rows)
    df = df.sort_values('Frequency', ascending=False)
    return df


def generate_cooccurrence_matrix(
    coding_results: List[CodingResult],
    codes: List[str]
) -> pd.DataFrame:
    """
    Generate code co-occurrence matrix.
    
    Args:
        coding_results: List of coding results.
        codes: List of code IDs.
    
    Returns:
        DataFrame with co-occurrence counts.
    """
    matrix = {c1: {c2: 0 for c2 in codes} for c1 in codes}
    
    for result in coding_results:
        assigned = result.assigned_codes
        for i, code1 in enumerate(assigned):
            for code2 in assigned[i:]:
                if code1 in matrix and code2 in matrix[code1]:
                    matrix[code1][code2] += 1
                    if code1 != code2:
                        matrix[code2][code1] += 1
    
    return pd.DataFrame(matrix)


def write_codebook_to_excel(
    codebook: Codebook,
    worksheet: Worksheet,
    start_row: int = 1
) -> int:
    """
    Write codebook to Excel worksheet.
    
    Args:
        codebook: Codebook to write.
        worksheet: Target worksheet.
        start_row: Starting row.
    
    Returns:
        Next available row.
    """
    ws = worksheet
    row = start_row
    
    ws.cell(row=row, column=1, value=f"CODEBOOK: {codebook.name}")
    row += 2
    
    headers = ["Code ID", "Code Name", "Definition", "Examples", "Parent Code", "Theme"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    row += 1
    
    code_to_theme = {}
    for theme_id, theme in codebook.themes.items():
        for code_id in theme.codes:
            code_to_theme[code_id] = theme.name
    
    for code_id, code in codebook.codes.items():
        ws.cell(row=row, column=1, value=code.id)
        ws.cell(row=row, column=2, value=code.name)
        ws.cell(row=row, column=3, value=code.definition)
        ws.cell(row=row, column=4, value="; ".join(code.examples[:3]))
        ws.cell(row=row, column=5, value=code.parent_code or "")
        ws.cell(row=row, column=6, value=code_to_theme.get(code_id, ""))
        row += 1
    
    row += 2
    ws.cell(row=row, column=1, value="THEMES")
    row += 1
    
    theme_headers = ["Theme ID", "Theme Name", "Description", "Number of Codes"]
    for col, header in enumerate(theme_headers, 1):
        ws.cell(row=row, column=col, value=header)
    row += 1
    
    for theme_id, theme in codebook.themes.items():
        ws.cell(row=row, column=1, value=theme.id)
        ws.cell(row=row, column=2, value=theme.name)
        ws.cell(row=row, column=3, value=theme.description)
        ws.cell(row=row, column=4, value=len(theme.codes))
        row += 1
    
    return row + 1


def write_coding_results_to_excel(
    results: List[CodingResult],
    worksheet: Worksheet,
    start_row: int = 1
) -> int:
    """
    Write coding results to Excel.
    
    Args:
        results: Coding results to write.
        worksheet: Target worksheet.
        start_row: Starting row.
    
    Returns:
        Next available row.
    """
    ws = worksheet
    row = start_row
    
    headers = ["Response ID", "Response Text", "Assigned Codes", "Coder", "Confidence"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    row += 1
    
    for result in results:
        ws.cell(row=row, column=1, value=result.response_id)
        ws.cell(row=row, column=2, value=result.response_text[:200])
        ws.cell(row=row, column=3, value=", ".join(result.assigned_codes))
        ws.cell(row=row, column=4, value=result.coder_id)
        ws.cell(row=row, column=5, value=round(result.confidence, 2))
        row += 1
    
    return row + 1


def create_default_codebook_from_responses(
    responses: List[str],
    codebook_name: str = "Auto-Generated Codebook"
) -> Codebook:
    """
    Create a basic codebook from response text using keyword extraction.
    
    Args:
        responses: List of text responses.
        codebook_name: Name for the codebook.
    
    Returns:
        Generated Codebook.
    """
    word_freq = Counter()
    
    for response in responses:
        if not response or str(response).lower() == 'nan':
            continue
        words = re.findall(r'\b[a-zA-Z]{4,}\b', str(response).lower())
        word_freq.update(words)
    
    stop_words = {
        'that', 'this', 'with', 'from', 'have', 'been', 'were', 'they',
        'their', 'would', 'could', 'should', 'about', 'which', 'there',
        'what', 'when', 'where', 'will', 'very', 'just', 'more', 'some',
        'also', 'than', 'then', 'into', 'only', 'other', 'such', 'these'
    }
    
    top_words = [
        word for word, count in word_freq.most_common(50)
        if word not in stop_words and count >= 3
    ][:20]
    
    codebook = Codebook(name=codebook_name)
    
    for i, word in enumerate(top_words, 1):
        code_id = f"C{i:02d}"
        code = Code(
            id=code_id,
            name=word.capitalize(),
            definition=f"Responses mentioning or relating to '{word}'",
            examples=[word]
        )
        codebook.add_code(code)
    
    if len(top_words) >= 5:
        theme1 = Theme(
            id="T01",
            name="Primary Themes",
            description="Most frequently occurring concepts",
            codes=[f"C{i:02d}" for i in range(1, min(6, len(top_words) + 1))]
        )
        codebook.add_theme(theme1)
    
    if len(top_words) >= 10:
        theme2 = Theme(
            id="T02",
            name="Secondary Themes",
            description="Additional recurring concepts",
            codes=[f"C{i:02d}" for i in range(6, min(11, len(top_words) + 1))]
        )
        codebook.add_theme(theme2)
    
    return codebook
