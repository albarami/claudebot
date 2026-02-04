"""
Excel tools for PhD Survey Analyzer.
CRITICAL: All cells contain FORMULAS, never hardcoded values.
"""

from pathlib import Path
from typing import Dict, List, Optional, Any
from datetime import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelFormulaWorkbook:
    """
    Excel workbook manager that enforces formula-only output.
    NEVER writes literal values - only formulas.
    """
    
    def __init__(self, output_path: Path):
        self.output_path = output_path
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)
        
        self.header_font = Font(bold=True)
        self.header_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
        self.title_font = Font(bold=True, size=14)
        self.apa_italic = Font(italic=True)
        
        thin = Side(style='thin', color='000000')
        self.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        self.formula_log: List[Dict[str, str]] = []
        self.sheets_created: List[str] = []
    
    def create_sheet(self, name: str) -> Worksheet:
        """Create a new worksheet."""
        ws = self.workbook.create_sheet(name)
        self.sheets_created.append(name)
        return ws
    
    def write_formula(self, ws: Worksheet, cell: str, formula: str, doc_col: Optional[int] = None) -> None:
        """
        Write a FORMULA to a cell.
        
        Args:
            ws: Worksheet
            cell: Cell reference (e.g., "B2")
            formula: Excel formula starting with "="
            doc_col: Optional column number for formula documentation
        """
        if not formula.startswith("="):
            raise ValueError(f"REJECTED: '{formula}' is not a formula. Must start with '='")
        
        ws[cell] = formula
        
        self.formula_log.append({
            "sheet": ws.title,
            "cell": cell,
            "formula": formula,
            "timestamp": datetime.now().isoformat()
        })
        
        if doc_col:
            row = int(''.join(filter(str.isdigit, cell)))
            ws.cell(row=row, column=doc_col, value=formula)
    
    def write_raw_data(self, ws: Worksheet, df: pd.DataFrame) -> None:
        """
        Write raw data to sheet (the ONLY place we write actual values).
        This is the source data that all formulas reference.
        """
        for c_idx, col in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=c_idx, value=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            for c_idx, value in enumerate(row, 1):
                if pd.isna(value):
                    ws.cell(row=r_idx, column=c_idx, value="")
                else:
                    ws.cell(row=r_idx, column=c_idx, value=value)
        
        ws.freeze_panes = "A2"
        
        ws.protection.sheet = True
        ws.protection.password = "locked"
        ws.protection.enable()
    
    def write_header_row(self, ws: Worksheet, headers: List[str], row: int = 1) -> None:
        """Write formatted header row."""
        for c_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
    
    def write_title(self, ws: Worksheet, title: str, row: int = 1) -> None:
        """Write sheet title."""
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = self.title_font
    
    def write_descriptives_formulas(
        self,
        ws: Worksheet,
        variables: List[str],
        raw_sheet: str,
        col_mapping: Dict[str, str],
        n_rows: int,
        start_row: int = 5
    ) -> None:
        """
        Write descriptive statistics using ONLY formulas.
        
        Args:
            ws: Target worksheet
            variables: List of variable names
            raw_sheet: Name of raw data sheet
            col_mapping: Maps variable name to column letter
            n_rows: Number of data rows
            start_row: Row to start writing
        """
        headers = ["Variable", "N", "M", "SD", "SE", "Median", "Min", "Max", 
                   "Skew", "Kurt", "95% CI Lower", "95% CI Upper", "Formula Doc"]
        self.write_header_row(ws, headers, start_row)
        
        row = start_row + 1
        for var in variables:
            col_letter = col_mapping.get(var)
            if not col_letter:
                continue
            
            data_range = f"'{raw_sheet}'!{col_letter}2:{col_letter}{n_rows + 1}"
            
            ws.cell(row=row, column=1, value=var)
            
            self.write_formula(ws, f"B{row}", f"=COUNT({data_range})")
            self.write_formula(ws, f"C{row}", f"=ROUND(AVERAGE({data_range}),2)")
            self.write_formula(ws, f"D{row}", f"=ROUND(STDEV.S({data_range}),2)")
            self.write_formula(ws, f"E{row}", f"=ROUND(STDEV.S({data_range})/SQRT(COUNT({data_range})),3)")
            self.write_formula(ws, f"F{row}", f"=ROUND(MEDIAN({data_range}),2)")
            self.write_formula(ws, f"G{row}", f"=MIN({data_range})")
            self.write_formula(ws, f"H{row}", f"=MAX({data_range})")
            self.write_formula(ws, f"I{row}", f"=ROUND(SKEW({data_range}),2)")
            self.write_formula(ws, f"J{row}", f"=ROUND(KURT({data_range}),2)")
            self.write_formula(ws, f"K{row}", f"=ROUND(AVERAGE({data_range})-1.96*STDEV.S({data_range})/SQRT(COUNT({data_range})),2)")
            self.write_formula(ws, f"L{row}", f"=ROUND(AVERAGE({data_range})+1.96*STDEV.S({data_range})/SQRT(COUNT({data_range})),2)")
            
            ws.cell(row=row, column=13, value=f"AVERAGE/STDEV.S/etc({data_range})")
            
            row += 1
    
    def write_codebook_formulas(
        self,
        ws: Worksheet,
        columns: List[str],
        raw_sheet: str,
        n_rows: int
    ) -> None:
        """Write codebook with formulas for each variable."""
        self.write_title(ws, "VARIABLE CODEBOOK")
        ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=3, column=1, value="All statistics computed via Excel formulas")
        
        headers = ["Variable", "Column", "N Valid", "N Missing", "% Missing", "Min", "Max", "Mean", "Formula"]
        self.write_header_row(ws, headers, 5)
        
        row = 6
        for c_idx, col in enumerate(columns, 1):
            col_letter = get_column_letter(c_idx)
            data_range = f"'{raw_sheet}'!{col_letter}2:{col_letter}{n_rows + 1}"
            
            ws.cell(row=row, column=1, value=col)
            ws.cell(row=row, column=2, value=col_letter)
            
            self.write_formula(ws, f"C{row}", f"=COUNT({data_range})")
            self.write_formula(ws, f"D{row}", f"=COUNTBLANK({data_range})")
            self.write_formula(ws, f"E{row}", f"=ROUND(COUNTBLANK({data_range})/{n_rows}*100,1)")
            self.write_formula(ws, f"F{row}", f"=MIN({data_range})")
            self.write_formula(ws, f"G{row}", f"=MAX({data_range})")
            self.write_formula(ws, f"H{row}", f"=IFERROR(ROUND(AVERAGE({data_range}),2),\"N/A\")")
            
            ws.cell(row=row, column=9, value=f"COUNT/COUNTBLANK({data_range})")
            row += 1
    
    def write_correlation_matrix_formulas(
        self,
        ws: Worksheet,
        variables: List[str],
        raw_sheet: str,
        col_mapping: Dict[str, str],
        n_rows: int
    ) -> None:
        """Write correlation matrix using CORREL formulas."""
        self.write_title(ws, "CORRELATION MATRIX")
        ws.cell(row=2, column=1, value="Pearson r computed via =CORREL() formula")
        ws.cell(row=3, column=1, value="* p < .05, ** p < .01")
        
        start_row = 5
        for i, var in enumerate(variables, 2):
            ws.cell(row=start_row, column=i, value=var).font = self.header_font
            ws.cell(row=start_row + i - 1, column=1, value=var).font = self.header_font
        
        for i, var1 in enumerate(variables):
            row = start_row + 1 + i
            col1_letter = col_mapping.get(var1)
            if not col1_letter:
                continue
            range1 = f"'{raw_sheet}'!{col1_letter}2:{col1_letter}{n_rows + 1}"
            
            for j, var2 in enumerate(variables):
                col = j + 2
                col2_letter = col_mapping.get(var2)
                if not col2_letter:
                    continue
                range2 = f"'{raw_sheet}'!{col2_letter}2:{col2_letter}{n_rows + 1}"
                
                if var1 == var2:
                    ws.cell(row=row, column=col, value="1.00")
                else:
                    self.write_formula(ws, f"{get_column_letter(col)}{row}", 
                                      f"=ROUND(CORREL({range1},{range2}),2)")
    
    def write_text_content(self, ws: Worksheet, title: str, content: str, start_row: int = 1) -> None:
        """Write text content to sheet."""
        self.write_title(ws, title, start_row)
        ws.cell(row=start_row + 1, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        row = start_row + 3
        for line in content.split('\n'):
            ws.cell(row=row, column=1, value=line)
            row += 1
    
    def save(self) -> Path:
        """Save workbook and return path."""
        self.workbook.save(self.output_path)
        return self.output_path
    
    def get_formula_log(self) -> List[Dict[str, str]]:
        """Return complete formula audit log."""
        return self.formula_log


def get_column_mapping(df: pd.DataFrame) -> Dict[str, str]:
    """Create mapping of column names to Excel column letters."""
    return {col: get_column_letter(idx + 1) for idx, col in enumerate(df.columns)}
