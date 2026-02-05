"""
Agent 2: Survey Implementer
Executes ONE task at a time by ACTUALLY creating Excel files with formulas.
Uses Claude Opus 4.5 for highest capability execution.
"""

from pathlib import Path
from typing import Dict, Any, List
from datetime import datetime
import json
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from langchain_anthropic import ChatAnthropic
from langchain_core.messages import HumanMessage, SystemMessage

from config import IMPLEMENTER_MODEL, IMPLEMENTER_TEMP, IMPLEMENTER_MAX_TOKENS, ANTHROPIC_API_KEY, OUTPUT_DIR
from utils.prompts import IMPLEMENTER_SYSTEM_PROMPT
from graph.state import SurveyAnalysisState, LogEntry
from tools.excel_tools import ExcelFormulaWorkbook, get_column_mapping


def execute_task_in_excel(
    task: Dict,
    df: pd.DataFrame,
    workbook_path: Path,
    col_mapping: Dict[str, str],
    n_rows: int
) -> Dict[str, Any]:
    """
    ACTUALLY execute the task by writing to Excel file.
    
    Args:
        task: Task specification
        df: Raw data DataFrame
        workbook_path: Path to output workbook
        col_mapping: Column name to letter mapping
        n_rows: Number of data rows
    
    Returns:
        Execution result with formulas written
    """
    if workbook_path.exists():
        wb = load_workbook(workbook_path)
    else:
        wb_manager = ExcelFormulaWorkbook(workbook_path)
        raw_ws = wb_manager.create_sheet("00_RAW_DATA_LOCKED")
        wb_manager.write_raw_data(raw_ws, df)
        wb_manager.save()
        wb = load_workbook(workbook_path)
    
    # Generate sheet name - extract from task's output_sheet field
    raw_output = task.get('output_sheet', '').strip()
    
    # Extract sheet name from patterns like: Sheet "01_DATA_AUDIT" or "01_DATA_AUDIT"
    sheet_name_match = re.search(r'["\']([^"\']+)["\']', raw_output)
    if sheet_name_match:
        raw_sheet_name = sheet_name_match.group(1).strip()
    else:
        # Try to extract from patterns like: 01_DATA_AUDIT sheet or Sheet 01_DATA_AUDIT
        word_match = re.search(r'\b(\d+_[A-Z_]+)\b', raw_output)
        if word_match:
            raw_sheet_name = word_match.group(1)
        else:
            raw_sheet_name = ''
    
    # Fallback: use task name if no sheet name found
    if not raw_sheet_name:
        task_name_clean = task.get('name', '').strip()
        if task_name_clean:
            raw_sheet_name = re.sub(r'[^a-zA-Z0-9_\- ]', '', task_name_clean)[:25]
        if not raw_sheet_name:
            task_id = task.get('id', '1.0').replace('.', '_')
            phase = task.get('phase', 'General')[:10]
            raw_sheet_name = f"{task_id}_{phase}"
    
    # Sanitize for Excel (31 chars max, no special chars)
    sheet_name = re.sub(r'[\\/*?:\[\]\n\r]', '', raw_sheet_name)[:31]
    sheet_name = sheet_name.strip()
    if not sheet_name:
        sheet_name = f"Task_{task.get('id', '1')}"
    
    print(f"[IMPLEMENTER] output_sheet field: '{raw_output[:100]}...'")
    print(f"[IMPLEMENTER] Extracted sheet name: '{sheet_name}'")
    
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
    
    formulas_written = []
    task_phase = task.get('phase', '').lower()
    task_name = task.get('name', '').lower()
    task_objective = task.get('objective', '').lower()
    task_method = task.get('method', '').lower()
    
    # Combine all task text for better matching
    task_text = f"{task_name} {task_phase} {task_objective} {task_method}"
    
    # Check for comprehensive mode (triggered by QC feedback)
    require_comprehensive = task.get('require_comprehensive', False)
    revision_number = task.get('revision_number', 0)
    
    # Expand column limit if comprehensive mode or revision
    col_limit = 50 if require_comprehensive or revision_number > 0 else 30
    numeric_cols = [col for col in df.columns if pd.api.types.is_numeric_dtype(df[col])][:col_limit]
    
    ws.cell(row=1, column=1, value=f"Task: {task['name']}")
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws.cell(row=3, column=1, value="All values computed via Excel formulas")
    
    # Match task type from combined task text
    if any(kw in task_text for kw in ['codebook', 'dictionary', 'variable', 'document']):
        # PhD-level Data Dictionary with proper academic structure
        ws.cell(row=5, column=1, value="=== VARIABLE METADATA ===")
        headers = [
            "Variable Name", "Column", "Data Type", "Measurement Level", 
            "N Valid", "N Missing", "% Complete", "Min", "Max", 
            "Unique Values", "Mean/Mode", "SD", "Range"
        ]
        for c, h in enumerate(headers, 1):
            ws.cell(row=6, column=c, value=h)
        
        row = 7
        for col_name in df.columns[:col_limit]:
            col_letter = col_mapping.get(col_name)
            if not col_letter:
                continue
            data_range = f"'00_RAW_DATA_LOCKED'!{col_letter}2:{col_letter}{n_rows + 1}"
            
            # Determine data type and measurement level
            col_data = df[col_name]
            is_numeric = pd.api.types.is_numeric_dtype(col_data)
            unique_count = col_data.nunique()
            
            if is_numeric:
                if unique_count <= 2:
                    data_type, meas_level = "Binary", "Nominal"
                elif unique_count <= 7:
                    data_type, meas_level = "Ordinal", "Ordinal"
                else:
                    data_type, meas_level = "Continuous", "Interval/Ratio"
            else:
                data_type, meas_level = "Categorical", "Nominal"
            
            ws.cell(row=row, column=1, value=col_name)
            ws.cell(row=row, column=2, value=col_letter)
            ws.cell(row=row, column=3, value=data_type)
            ws.cell(row=row, column=4, value=meas_level)
            
            # Formulas for statistics - academically correct
            f_valid = f"=COUNTA({data_range})"  # Count all non-blank cells
            f_missing = f"=COUNTBLANK({data_range})"
            f_complete = f"=ROUND(COUNTA({data_range})/{n_rows}*100,1)"
            f_min = f"=IFERROR(MIN({data_range}),\"N/A\")"
            f_max = f"=IFERROR(MAX({data_range}),\"N/A\")"
            f_unique = f"=IFERROR(SUMPRODUCT(1/COUNTIF({data_range},{data_range})),0)"
            f_central = f"=IFERROR(ROUND(AVERAGE({data_range}),2),IFERROR(MODE.SNGL({data_range}),\"N/A\"))"
            f_sd = f"=IFERROR(ROUND(STDEV.S({data_range}),2),\"N/A\")"
            f_range = f"=IFERROR(MAX({data_range})-MIN({data_range}),\"N/A\")"
            
            ws.cell(row=row, column=5, value=f_valid)
            ws.cell(row=row, column=6, value=f_missing)
            ws.cell(row=row, column=7, value=f_complete)
            ws.cell(row=row, column=8, value=f_min)
            ws.cell(row=row, column=9, value=f_max)
            ws.cell(row=row, column=10, value=f_unique)
            ws.cell(row=row, column=11, value=f_central)
            ws.cell(row=row, column=12, value=f_sd)
            ws.cell(row=row, column=13, value=f_range)
            
            formulas_written.extend([
                {"cell": f"E{row}", "formula": f_valid, "purpose": "N valid responses"},
                {"cell": f"F{row}", "formula": f_missing, "purpose": "N missing"},
                {"cell": f"G{row}", "formula": f_complete, "purpose": "% completeness"},
                {"cell": f"H{row}", "formula": f_min, "purpose": "Minimum value"},
                {"cell": f"I{row}", "formula": f_max, "purpose": "Maximum value"},
                {"cell": f"J{row}", "formula": f_unique, "purpose": "Unique value count"},
                {"cell": f"K{row}", "formula": f_central, "purpose": "Central tendency"},
                {"cell": f"L{row}", "formula": f_sd, "purpose": "Standard deviation"},
                {"cell": f"M{row}", "formula": f_range, "purpose": "Range"}
            ])
            row += 1
        
        # Add summary section
        summary_row = row + 2
        ws.cell(row=summary_row, column=1, value="=== DATASET SUMMARY ===")
        ws.cell(row=summary_row+1, column=1, value="Total Variables:")
        ws.cell(row=summary_row+1, column=2, value=f"={len(df.columns)}")
        ws.cell(row=summary_row+2, column=1, value="Total Observations:")
        ws.cell(row=summary_row+2, column=2, value=f"={n_rows}")
        ws.cell(row=summary_row+3, column=1, value="Numeric Variables:")
        ws.cell(row=summary_row+3, column=2, value=f"={len(numeric_cols)}")
        formulas_written.extend([
            {"cell": f"B{summary_row+1}", "formula": f"={len(df.columns)}", "purpose": "Variable count"},
            {"cell": f"B{summary_row+2}", "formula": f"={n_rows}", "purpose": "Observation count"}
        ])
    
    elif any(kw in task_text for kw in ['descriptive', 'statistics', 'mean', 'average', 'summary']):
        headers = ["Variable", "N", "Mean", "SD", "SE", "Median", "Min", "Max", "Skew", "Kurt"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=5, column=c, value=h)
        
        row = 6
        for col_name in numeric_cols:
            col_letter = col_mapping.get(col_name)
            if not col_letter:
                continue
            data_range = f"'00_RAW_DATA_LOCKED'!{col_letter}2:{col_letter}{n_rows + 1}"
            
            ws.cell(row=row, column=1, value=col_name)
            
            formulas = [
                (2, f"=COUNT({data_range})", "N"),
                (3, f"=ROUND(AVERAGE({data_range}),2)", "Mean"),
                (4, f"=ROUND(STDEV.S({data_range}),2)", "SD"),
                (5, f"=ROUND(STDEV.S({data_range})/SQRT(COUNT({data_range})),3)", "SE"),
                (6, f"=ROUND(MEDIAN({data_range}),2)", "Median"),
                (7, f"=MIN({data_range})", "Min"),
                (8, f"=MAX({data_range})", "Max"),
                (9, f"=ROUND(SKEW({data_range}),2)", "Skew"),
                (10, f"=ROUND(KURT({data_range}),2)", "Kurtosis")
            ]
            
            for col_idx, formula, purpose in formulas:
                ws.cell(row=row, column=col_idx, value=formula)
                formulas_written.append({
                    "cell": f"{get_column_letter(col_idx)}{row}",
                    "formula": formula,
                    "purpose": purpose
                })
            row += 1
    
    elif any(kw in task_text for kw in ['correlation', 'correl', 'relationship', 'association']):
        vars_to_correlate = numeric_cols[:10]
        
        for i, var in enumerate(vars_to_correlate, 2):
            ws.cell(row=5, column=i, value=var)
            ws.cell(row=5 + i - 1, column=1, value=var)
        
        for i, var1 in enumerate(vars_to_correlate):
            row = 6 + i
            col1_letter = col_mapping.get(var1)
            if not col1_letter:
                continue
            range1 = f"'00_RAW_DATA_LOCKED'!{col1_letter}2:{col1_letter}{n_rows + 1}"
            
            for j, var2 in enumerate(vars_to_correlate):
                col = j + 2
                col2_letter = col_mapping.get(var2)
                if not col2_letter:
                    continue
                range2 = f"'00_RAW_DATA_LOCKED'!{col2_letter}2:{col2_letter}{n_rows + 1}"
                
                if var1 == var2:
                    ws.cell(row=row, column=col, value="1.00")
                else:
                    formula = f"=ROUND(CORREL({range1},{range2}),2)"
                    ws.cell(row=row, column=col, value=formula)
                    formulas_written.append({
                        "cell": f"{get_column_letter(col)}{row}",
                        "formula": formula,
                        "purpose": f"Correlation {var1} x {var2}"
                    })
    
    elif any(kw in task_text for kw in ['quality', 'missing', 'complete', 'clean', 'valid', 'integrity', 'check']):
        # Data Quality / Integrity Analysis - comprehensive metrics
        headers = ["Variable", "N Valid", "N Missing", "% Complete", "Min", "Max", "Range", "Mean", "SD"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=5, column=c, value=h)
        
        # Summary metrics at top
        total_cells = len(df.columns) * n_rows
        ws.cell(row=6, column=1, value="=== DATASET SUMMARY ===")
        ws.cell(row=7, column=1, value="Total Variables")
        ws.cell(row=7, column=2, value=f"={len(df.columns)}")
        ws.cell(row=8, column=1, value="Total Observations")
        ws.cell(row=8, column=2, value=f"={n_rows}")
        ws.cell(row=9, column=1, value="Total Cells")
        ws.cell(row=9, column=2, value=f"={total_cells}")
        
        formulas_written.extend([
            {"cell": "B7", "formula": f"={len(df.columns)}", "purpose": "Total vars"},
            {"cell": "B8", "formula": f"={n_rows}", "purpose": "Total obs"},
            {"cell": "B9", "formula": f"={total_cells}", "purpose": "Total cells"}
        ])
        
        # Per-variable quality metrics
        ws.cell(row=11, column=1, value="=== PER-VARIABLE QUALITY ===")
        for c, h in enumerate(headers, 1):
            ws.cell(row=12, column=c, value=h)
        
        row = 13
        for col_name in df.columns[:col_limit]:
            col_letter = col_mapping.get(col_name)
            if not col_letter:
                continue
            data_range = f"'00_RAW_DATA_LOCKED'!{col_letter}2:{col_letter}{n_rows + 1}"
            
            ws.cell(row=row, column=1, value=col_name)
            
            f1 = f"=COUNT({data_range})"
            f2 = f"=COUNTBLANK({data_range})"
            f3 = f"=ROUND((COUNT({data_range})/{n_rows})*100,1)"
            f4 = f"=MIN({data_range})"
            f5 = f"=MAX({data_range})"
            f6 = f"=MAX({data_range})-MIN({data_range})"
            f7 = f"=IFERROR(ROUND(AVERAGE({data_range}),2),\"N/A\")"
            f8 = f"=IFERROR(ROUND(STDEV.S({data_range}),2),\"N/A\")"
            
            ws.cell(row=row, column=2, value=f1)
            ws.cell(row=row, column=3, value=f2)
            ws.cell(row=row, column=4, value=f3)
            ws.cell(row=row, column=5, value=f4)
            ws.cell(row=row, column=6, value=f5)
            ws.cell(row=row, column=7, value=f6)
            ws.cell(row=row, column=8, value=f7)
            ws.cell(row=row, column=9, value=f8)
            
            formulas_written.extend([
                {"cell": f"B{row}", "formula": f1, "purpose": f"N valid for {col_name}"},
                {"cell": f"C{row}", "formula": f2, "purpose": f"N missing for {col_name}"},
                {"cell": f"D{row}", "formula": f3, "purpose": f"% complete for {col_name}"},
                {"cell": f"E{row}", "formula": f4, "purpose": f"Min for {col_name}"},
                {"cell": f"H{row}", "formula": f7, "purpose": f"Mean for {col_name}"}
            ])
            row += 1
    
    else:
        # Default: Create a general analysis sheet with multiple formulas
        ws.cell(row=5, column=1, value="Phase:")
        ws.cell(row=5, column=2, value=task.get('phase', 'General'))
        ws.cell(row=6, column=1, value="Objective:")
        ws.cell(row=6, column=2, value=task.get('objective', 'N/A')[:100])
        
        # Create summary statistics for first 5 numeric columns
        headers = ["Variable", "N", "Mean", "SD", "Min", "Max"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=8, column=c, value=h)
        
        # Use more columns in comprehensive mode
        cols_to_process = numeric_cols[:15] if (require_comprehensive or revision_number > 0) else numeric_cols[:5]
        row = 9
        for col_name in cols_to_process:
            col_letter = col_mapping.get(col_name)
            if not col_letter:
                continue
            data_range = f"'00_RAW_DATA_LOCKED'!{col_letter}2:{col_letter}{n_rows + 1}"
            
            ws.cell(row=row, column=1, value=col_name)
            
            f1 = f"=COUNT({data_range})"
            f2 = f"=ROUND(AVERAGE({data_range}),2)"
            f3 = f"=ROUND(STDEV.S({data_range}),2)"
            f4 = f"=MIN({data_range})"
            f5 = f"=MAX({data_range})"
            
            ws.cell(row=row, column=2, value=f1)
            ws.cell(row=row, column=3, value=f2)
            ws.cell(row=row, column=4, value=f3)
            ws.cell(row=row, column=5, value=f4)
            ws.cell(row=row, column=6, value=f5)
            
            formulas_written.extend([
                {"cell": f"B{row}", "formula": f1, "purpose": f"N for {col_name}"},
                {"cell": f"C{row}", "formula": f2, "purpose": f"Mean for {col_name}"},
                {"cell": f"D{row}", "formula": f3, "purpose": f"SD for {col_name}"},
                {"cell": f"E{row}", "formula": f4, "purpose": f"Min for {col_name}"},
                {"cell": f"F{row}", "formula": f5, "purpose": f"Max for {col_name}"}
            ])
            row += 1
    
    wb.save(workbook_path)
    
    return {
        "sheet_created": sheet_name,
        "formulas_written": formulas_written,
        "workbook_path": str(workbook_path)
    }


async def implementer_node(state: SurveyAnalysisState) -> Dict[str, Any]:
    """
    Implementer agent node - ACTUALLY executes task in Excel.
    Reads QC feedback and adjusts execution accordingly.
    
    Args:
        state: Current workflow state
    
    Returns:
        State updates with task output and Excel modifications
    """
    current_idx = state['current_task_idx']
    tasks = state['tasks']
    revision_count = state.get('task_revision_count', 0)
    qc_feedback = state.get('qc_feedback', '')
    
    if current_idx >= len(tasks):
        return {
            "status": "all_tasks_complete",
            "messages": [{"role": "implementer", "content": "All tasks completed"}]
        }
    
    current_task = tasks[current_idx]
    
    # If this is a revision, enhance the task based on QC feedback
    if revision_count > 0 and qc_feedback:
        # Parse feedback and enhance task requirements
        enhanced_task = dict(current_task)
        enhanced_task['qc_feedback'] = qc_feedback
        enhanced_task['revision_number'] = revision_count
        
        # Add more comprehensive requirements based on feedback patterns
        feedback_lower = qc_feedback.lower()
        if 'formula' in feedback_lower or 'more' in feedback_lower:
            enhanced_task['require_comprehensive'] = True
        if 'verify' in feedback_lower or 'check' in feedback_lower:
            enhanced_task['require_validation'] = True
        if 'document' in feedback_lower or 'explain' in feedback_lower:
            enhanced_task['require_documentation'] = True
        
        current_task = enhanced_task
    
    file_path = Path(state['file_path'])
    df = pd.read_excel(file_path)
    col_mapping = get_column_mapping(df)
    n_rows = len(df)
    
    session_id = state['session_id']
    workbook_path = OUTPUT_DIR / f"PhD_EDA_{session_id}.xlsx"
    
    excel_result = execute_task_in_excel(
        task=current_task,
        df=df,
        workbook_path=workbook_path,
        col_mapping=col_mapping,
        n_rows=n_rows
    )
    
    sheets_created = state.get('sheets_created', [])
    sheets_created.append(excel_result['sheet_created'])
    
    formulas_documented = state.get('formulas_documented', [])
    formulas_documented.extend(excel_result['formulas_written'])
    
    task_output = f"""
TASK COMPLETED: {current_task['id']} - {current_task['name']}

EXCEL FILE: {excel_result['workbook_path']}
SHEET CREATED: {excel_result['sheet_created']}

FORMULAS WRITTEN ({len(excel_result['formulas_written'])}):
| Cell | Formula | Purpose |
|------|---------|---------|
"""
    for f in excel_result['formulas_written'][:20]:
        task_output += f"| {f['cell']} | {f['formula'][:50]}... | {f['purpose']} |\n"
    
    if len(excel_result['formulas_written']) > 20:
        task_output += f"... and {len(excel_result['formulas_written']) - 20} more formulas\n"
    
    task_output += "\nREADY FOR QC REVIEW - Please verify the Excel file directly."
    
    updated_task = dict(current_task)
    updated_task['status'] = 'in_review'
    
    updated_tasks = list(tasks)
    updated_tasks[current_idx] = updated_task
    
    log_entry = LogEntry(
        timestamp=datetime.now().isoformat(),
        agent="implementer",
        action=f"Executed task {current_task['id']} in Excel",
        details=f"Created sheet: {excel_result['sheet_created']}, Formulas: {len(excel_result['formulas_written'])}",
        task_id=current_task['id']
    )
    
    return {
        "current_task": updated_task,
        "current_task_output": task_output,
        "tasks": updated_tasks,
        "sheets_created": sheets_created,
        "formulas_documented": formulas_documented,
        "workbook_path": str(workbook_path),
        "status": "reviewing",
        "execution_log": [log_entry],
        "messages": [{"role": "implementer", "content": f"Created sheet '{excel_result['sheet_created']}' with {len(excel_result['formulas_written'])} formulas"}]
    }
