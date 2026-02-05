"""
Agent 3: QC Reviewer - DUAL REVIEW SYSTEM
Verifies every task meets PhD-level standards with VETO POWER.

Architecture:
1. DETERMINISTIC QC: Fast, programmatic checks (formula coverage, references)
2. LLM QC: Dual review by Claude Sonnet 4.5 AND OpenAI 5.2

Reviews the ACTUAL Excel file, not just text reports.
"""

from pathlib import Path
from typing import Dict, Any, List
from datetime import datetime
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from langchain_anthropic import ChatAnthropic
from langchain_openai import ChatOpenAI
from langchain_core.messages import HumanMessage, SystemMessage

from config import (
    QC_REVIEWER_MODEL_1, QC_REVIEWER_MODEL_2,
    QC_REVIEWER_TEMP, QC_REVIEWER_MAX_TOKENS,
    ANTHROPIC_API_KEY, OPENAI_API_KEY
)
from utils.prompts import QC_REVIEWER_SYSTEM_PROMPT
from graph.state import SurveyAnalysisState, QCDecision, LogEntry
from engines.qc_engine import run_deterministic_qc


def verify_excel_file(workbook_path: str, sheet_name: str) -> Dict[str, Any]:
    """
    ACTUALLY read and verify the Excel file.
    
    Args:
        workbook_path: Path to the Excel workbook
        sheet_name: Name of sheet to verify
    
    Returns:
        Verification results with formula counts, errors, etc.
    """
    result = {
        "file_exists": False,
        "sheet_exists": False,
        "total_cells": 0,
        "formula_cells": 0,
        "value_cells": 0,
        "empty_cells": 0,
        "formula_percentage": 0,
        "sample_formulas": [],
        "potential_errors": [],
        "cell_contents": []
    }
    
    path = Path(workbook_path)
    if not path.exists():
        result["potential_errors"].append(f"Excel file not found: {workbook_path}")
        return result
    
    result["file_exists"] = True
    
    try:
        wb = load_workbook(workbook_path, data_only=False)
        
        if sheet_name not in wb.sheetnames:
            result["potential_errors"].append(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            return result
        
        result["sheet_exists"] = True
        ws = wb[sheet_name]
        
        for row in ws.iter_rows(min_row=1, max_row=min(50, ws.max_row or 1), 
                                 min_col=1, max_col=min(15, ws.max_column or 1)):
            for cell in row:
                result["total_cells"] += 1
                
                if cell.value is None:
                    result["empty_cells"] += 1
                elif isinstance(cell.value, str) and cell.value.startswith("="):
                    result["formula_cells"] += 1
                    if len(result["sample_formulas"]) < 10:
                        result["sample_formulas"].append({
                            "cell": f"{get_column_letter(cell.column)}{cell.row}",
                            "formula": cell.value
                        })
                else:
                    result["value_cells"] += 1
                    result["cell_contents"].append({
                        "cell": f"{get_column_letter(cell.column)}{cell.row}",
                        "value": str(cell.value)[:50]
                    })
        
        non_empty = result["total_cells"] - result["empty_cells"]
        if non_empty > 0:
            result["formula_percentage"] = (result["formula_cells"] / non_empty) * 100
        
        if result["formula_percentage"] < 50 and result["formula_cells"] > 0:
            result["potential_errors"].append(
                f"Low formula percentage: {result['formula_percentage']:.1f}% - expected mostly formulas"
            )
        
        wb.close()
        
    except Exception as e:
        result["potential_errors"].append(f"Error reading Excel: {str(e)}")
    
    return result


def build_review_prompt(
    current_task: Dict, 
    task_output: str, 
    revision_count: int, 
    prev_feedback: str,
    excel_verification: Dict[str, Any]
) -> str:
    """Build the review prompt with actual Excel verification data."""
    
    excel_report = f"""
EXCEL FILE VERIFICATION (ACTUAL FILE CONTENTS):
- File exists: {excel_verification['file_exists']}
- Sheet exists: {excel_verification['sheet_exists']}
- Total cells examined: {excel_verification['total_cells']}
- Cells with formulas: {excel_verification['formula_cells']}
- Cells with values: {excel_verification['value_cells']}
- Formula percentage: {excel_verification['formula_percentage']:.1f}%

SAMPLE FORMULAS FOUND IN EXCEL:
"""
    for f in excel_verification.get('sample_formulas', [])[:5]:
        excel_report += f"  {f['cell']}: {f['formula']}\n"
    
    if excel_verification.get('potential_errors'):
        excel_report += "\nPOTENTIAL ISSUES:\n"
        for err in excel_verification['potential_errors']:
            excel_report += f"  ⚠️ {err}\n"
    
    return f"""Review this task execution for PhD-level quality:

TASK SPECIFICATION (from Master Plan):
- Task ID: {current_task['id']}
- Phase: {current_task['phase']}
- Name: {current_task['name']}
- Objective: {current_task['objective']}
- Method: {current_task['method']}
- Expected Output: {current_task['output_sheet']}

{excel_report}

IMPLEMENTER'S REPORT:
{task_output}

REVISION HISTORY:
This is revision attempt #{revision_count + 1}
{f"Previous QC feedback: {prev_feedback}" if revision_count > 0 else "First submission"}

YOUR VERIFICATION CHECKLIST:

A. EXCEL FILE VERIFICATION (from actual file inspection above)
☐ Excel file exists and is accessible
☐ Required sheet was created
☐ Cells contain formulas (check formula percentage above)
☐ Formulas use correct syntax (=AVERAGE, =STDEV.S, =CORREL, etc.)
☐ Formulas reference '00_RAW_DATA_LOCKED' sheet correctly

B. METHODOLOGICAL SOUNDNESS
☐ Statistical method matches the task objective
☐ Appropriate for the data type
☐ Sample size considerations addressed

C. FORMULA ACCURACY (verify from sample formulas above)
☐ Formulas would produce correct results
☐ Data ranges are appropriate
☐ No obvious errors in formula logic

DECISION CRITERIA:
- APPROVE: File exists, sheet created, formulas present and correct
- REJECT: Missing file/sheet, no formulas, or incorrect formulas
- CONDITIONAL: Minor issues that don't affect accuracy
- HALT: Only if task is fundamentally impossible

Make your decision: APPROVE, REJECT, CONDITIONAL, or HALT"""


def parse_decision(review_text: str) -> str:
    """Parse decision from review text."""
    if "REJECT" in review_text.upper():
        return "REJECT"
    elif "HALT" in review_text.upper():
        return "HALT"
    elif "CONDITIONAL" in review_text.upper():
        return "CONDITIONAL"
    return "APPROVE"


async def qc_reviewer_node(state: SurveyAnalysisState) -> Dict[str, Any]:
    """
    DUAL QC Reviewer - uses BOTH Sonnet 4.5 AND OpenAI 5.2.
    Both reviewers must agree for approval.
    If EITHER rejects, the task is rejected.
    
    Args:
        state: Current workflow state
    
    Returns:
        State updates with QC decision
    """
    current_task = state.get('current_task')
    task_output = state.get('current_task_output', '')
    
    if not current_task:
        return {
            "qc_decision": "ERROR",
            "qc_feedback": "No task to review",
            "messages": [{"role": "qc_reviewer", "content": "Error: No task to review"}]
        }
    
    revision_count = state.get('task_revision_count', 0)
    prev_feedback = state.get('qc_feedback', '')
    
    # ACTUALLY verify the Excel file
    workbook_path = state.get('workbook_path', '')
    
    # Get the sheet name that was actually created by the implementer
    sheets_created = state.get('sheets_created', [])
    if sheets_created:
        sheet_name = sheets_created[-1]  # Use the most recently created sheet
    else:
        # Fallback: use same logic as implementer (with markdown cleanup)
        raw_sheet_name = current_task.get('output_sheet', '').strip()
        if raw_sheet_name:
            raw_sheet_name = re.sub(r'^[-*•]\s*', '', raw_sheet_name)
            raw_sheet_name = re.sub(r'Sheet\s*["\']?', '', raw_sheet_name, flags=re.IGNORECASE)
            raw_sheet_name = re.sub(r'["\'\n\r]', '', raw_sheet_name)
            raw_sheet_name = raw_sheet_name.strip(' -')
        if not raw_sheet_name:
            task_id = current_task.get('id', '1.0').replace('.', '_')
            phase = current_task.get('phase', 'General')[:10]
            raw_sheet_name = f"{task_id}_{phase}"
        sheet_name = re.sub(r'[\\/*?:\[\]\n\r]', '', raw_sheet_name)[:31]
        sheet_name = sheet_name.strip()
        if not sheet_name:
            sheet_name = f"Task_{current_task.get('id', '1')}"
    
    excel_verification = verify_excel_file(workbook_path, sheet_name)
    
    # === STEP 1: DETERMINISTIC QC (fast, programmatic checks) ===
    deterministic_result = run_deterministic_qc(Path(workbook_path), sheet_name)
    print(deterministic_result.get("summary", ""))
    
    # If deterministic QC fails on critical checks, reject immediately (no LLM needed)
    if not deterministic_result.get("passed", False):
        deterministic_errors = deterministic_result.get("errors", [])
        if any("not found" in e.lower() or "coverage" in e.lower() for e in deterministic_errors):
            combined_feedback = f"""
DETERMINISTIC QC FAILED (no LLM review needed)

Errors:
{chr(10).join('- ' + e for e in deterministic_errors)}

Metrics:
- Formula coverage: {deterministic_result.get('metrics', {}).get('formula_percentage', 0)}%
- Formula cells: {deterministic_result.get('metrics', {}).get('formula_cells', 0)}

Fix these issues before resubmitting.
"""
            return {
                "qc_decision": "REJECT",
                "qc_feedback": combined_feedback,
                "task_revision_count": revision_count + 1,
                "execution_log": [LogEntry(
                    timestamp=datetime.now().isoformat(),
                    agent="qc_reviewer",
                    action=f"Deterministic QC failed for task {current_task['id']}",
                    details=f"Errors: {deterministic_errors}",
                    task_id=current_task['id']
                )],
                "messages": [
                    {"role": "qc_reviewer", "content": f"Deterministic QC: REJECT - {deterministic_errors[0] if deterministic_errors else 'Failed checks'}"}
                ]
            }
    
    # === STEP 2: LLM QC (methodology/quality review) ===
    prompt = build_review_prompt(current_task, task_output, revision_count, prev_feedback, excel_verification)
    messages = [
        SystemMessage(content=QC_REVIEWER_SYSTEM_PROMPT),
        HumanMessage(content=prompt)
    ]
    
    # === REVIEW 1: Claude Sonnet 4.5 ===
    llm_sonnet = ChatAnthropic(
        model=QC_REVIEWER_MODEL_1,
        temperature=QC_REVIEWER_TEMP,
        max_tokens=QC_REVIEWER_MAX_TOKENS,
        api_key=ANTHROPIC_API_KEY
    )
    
    response_sonnet = await llm_sonnet.ainvoke(messages)
    review_sonnet = response_sonnet.content
    decision_sonnet = parse_decision(review_sonnet)
    
    # Log Sonnet's decision for debugging
    print(f"\n{'='*60}")
    print(f"SONNET 4.5 REVIEW: {decision_sonnet}")
    print(f"{'='*60}")
    print(review_sonnet[:500] if len(review_sonnet) > 500 else review_sonnet)
    print(f"{'='*60}\n")
    
    # === REVIEW 2: OpenAI 5.2 ===
    llm_openai = ChatOpenAI(
        model=QC_REVIEWER_MODEL_2,
        temperature=QC_REVIEWER_TEMP,
        max_tokens=QC_REVIEWER_MAX_TOKENS,
        api_key=OPENAI_API_KEY
    )
    
    response_openai = await llm_openai.ainvoke(messages)
    review_openai = response_openai.content
    decision_openai = parse_decision(review_openai)
    
    # Log OpenAI's decision for debugging
    print(f"\n{'='*60}")
    print(f"OPENAI 5.2 REVIEW: {decision_openai}")
    print(f"{'='*60}")
    print(review_openai[:500] if len(review_openai) > 500 else review_openai)
    print(f"{'='*60}\n")
    
    # === DUAL REVIEW LOGIC ===
    # Real intelligent decision - no forced approvals
    # Both reviewers have equal weight
    
    if decision_sonnet == "HALT" or decision_openai == "HALT":
        final_decision = "HALT"
    elif decision_sonnet == "REJECT" or decision_openai == "REJECT":
        # If either rejects, we reject - but provide combined feedback
        final_decision = "REJECT"
    elif decision_sonnet == "APPROVE" and decision_openai == "APPROVE":
        final_decision = "APPROVE"
    elif decision_sonnet == "CONDITIONAL" or decision_openai == "CONDITIONAL":
        # If one approves and other is conditional, it's conditional
        final_decision = "CONDITIONAL"
    else:
        # Default to the more conservative decision
        final_decision = "APPROVE"
    
    # Combine feedback from both reviewers
    combined_feedback = f"""
═══════════════════════════════════════════════════════════
DUAL QC REVIEW RESULTS
═══════════════════════════════════════════════════════════

📋 REVIEW 1: Claude Sonnet 4.5
Decision: {decision_sonnet}
{review_sonnet}

───────────────────────────────────────────────────────────

📋 REVIEW 2: OpenAI 5.2
Decision: {decision_openai}
{review_openai}

───────────────────────────────────────────────────────────

🎯 FINAL DECISION: {final_decision}
(Both reviewers must agree for approval)
═══════════════════════════════════════════════════════════
"""
    
    qc_record = QCDecision(
        task_id=current_task['id'],
        decision=final_decision,
        feedback=combined_feedback,
        checklist_results={
            "sonnet_decision": decision_sonnet,
            "openai_decision": decision_openai
        },
        timestamp=datetime.now().isoformat(),
        revision_number=revision_count + 1
    )
    
    log_entry = LogEntry(
        timestamp=datetime.now().isoformat(),
        agent="qc_reviewer",
        action=f"Dual review of task {current_task['id']}",
        details=f"Sonnet: {decision_sonnet}, OpenAI: {decision_openai} → Final: {final_decision}",
        task_id=current_task['id']
    )
    
    # Only increment revision count on REJECT, reset on approval
    if final_decision == "REJECT":
        new_revision_count = revision_count + 1
    else:
        new_revision_count = 0
    
    return {
        "qc_decision": final_decision,
        "qc_feedback": combined_feedback,
        "qc_history": [qc_record],
        "task_revision_count": new_revision_count,
        "execution_log": [log_entry],
        "messages": [
            {"role": "qc_reviewer_sonnet", "content": f"Sonnet 4.5: {decision_sonnet}"},
            {"role": "qc_reviewer_openai", "content": f"OpenAI 5.2: {decision_openai}"},
            {"role": "qc_reviewer", "content": f"FINAL: {final_decision}"}
        ]
    }
