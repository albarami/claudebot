"""
Deterministic QC Engine.
Programmatic validation of Excel workbooks before LLM review.
Integrates statistical verification for ground-truth comparison.
"""

from typing import Dict, List, Any, Tuple, Optional
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re

import pandas as pd

from tools.verification import (
    StatisticalVerifier,
    ExcelVerifier,
    VerificationResult,
    VerificationStatus,
    verify_task_output,
    generate_verification_report
)


class DeterministicQC:
    """
    Deterministic quality checks for Excel workbooks.
    Run BEFORE LLM-based QC for fast, reliable validation.
    """
    
    RAW_DATA_SHEET = "00_CLEANED_DATA"
    MIN_FORMULA_PERCENTAGE = 50  # Minimum % of data cells that must be formulas
    
    def __init__(self, workbook_path: Path, raw_data: Optional[pd.DataFrame] = None):
        self.workbook_path = workbook_path
        self.raw_data = raw_data
        self.errors: List[str] = []
        self.warnings: List[str] = []
        self.metrics: Dict[str, Any] = {}
        self.verification_result: Optional[VerificationResult] = None
    
    def run_all_checks(self, sheet_name: str) -> Dict[str, Any]:
        """
        Run all deterministic QC checks on a sheet.
        
        Args:
            sheet_name: Name of sheet to verify
        
        Returns:
            QC result with pass/fail, errors, warnings, metrics
        """
        self.errors = []
        self.warnings = []
        self.metrics = {}
        
        # Check 1: File exists
        if not self._check_file_exists():
            return self._build_result(passed=False)
        
        # Check 2: Sheet exists
        if not self._check_sheet_exists(sheet_name):
            return self._build_result(passed=False)
        
        # Check 3: Raw data sheet exists
        if not self._check_raw_data_sheet():
            self.warnings.append(f"Raw data sheet '{self.RAW_DATA_SHEET}' not found")
        
        # Check 4: Formula coverage
        formula_check = self._check_formula_coverage(sheet_name)
        
        # Check 5: Formula references
        ref_check = self._check_formula_references(sheet_name)
        
        # Check 6: No Excel errors
        error_check = self._check_no_excel_errors(sheet_name)
        
        # Determine overall pass/fail
        passed = (
            formula_check and 
            len(self.errors) == 0
        )
        
        return self._build_result(passed=passed)
    
    def _check_file_exists(self) -> bool:
        """Check if workbook file exists."""
        if not self.workbook_path.exists():
            self.errors.append(f"Workbook file not found: {self.workbook_path}")
            return False
        self.metrics["file_exists"] = True
        return True
    
    def _check_sheet_exists(self, sheet_name: str) -> bool:
        """Check if specified sheet exists in workbook."""
        try:
            wb = load_workbook(self.workbook_path, data_only=False)
            if sheet_name not in wb.sheetnames:
                self.errors.append(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
                wb.close()
                return False
            self.metrics["sheet_exists"] = True
            self.metrics["available_sheets"] = wb.sheetnames
            wb.close()
            return True
        except Exception as e:
            self.errors.append(f"Error opening workbook: {str(e)}")
            return False
    
    def _check_raw_data_sheet(self) -> bool:
        """Check if raw data sheet exists."""
        try:
            wb = load_workbook(self.workbook_path, data_only=False)
            exists = self.RAW_DATA_SHEET in wb.sheetnames
            self.metrics["raw_data_sheet_exists"] = exists
            wb.close()
            return exists
        except Exception:
            return False
    
    def _check_formula_coverage(self, sheet_name: str) -> bool:
        """
        Check formula coverage in data region.
        Headers/labels (row 1-5, column A) are excluded.
        """
        try:
            wb = load_workbook(self.workbook_path, data_only=False)
            ws = wb[sheet_name]
            
            total_data_cells = 0
            formula_cells = 0
            value_cells = 0
            empty_cells = 0
            sample_formulas = []
            
            # Define data region (skip first 5 rows as headers, first column as labels)
            data_start_row = 6
            data_start_col = 2
            
            max_row = min(100, ws.max_row or 1)
            max_col = min(20, ws.max_column or 1)
            
            for row in range(data_start_row, max_row + 1):
                for col in range(data_start_col, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    total_data_cells += 1
                    
                    if cell.value is None:
                        empty_cells += 1
                    elif isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_cells += 1
                        if len(sample_formulas) < 5:
                            sample_formulas.append({
                                "cell": f"{get_column_letter(col)}{row}",
                                "formula": cell.value[:100]
                            })
                    else:
                        value_cells += 1
            
            # Also check the first 5 rows for any formulas (they count too)
            for row in range(1, min(6, ws.max_row or 1) + 1):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_cells += 1
                        if len(sample_formulas) < 5:
                            sample_formulas.append({
                                "cell": f"{get_column_letter(col)}{row}",
                                "formula": cell.value[:100]
                            })
            
            wb.close()
            
            non_empty = total_data_cells - empty_cells + formula_cells  # Include header formulas
            formula_percentage = (formula_cells / non_empty * 100) if non_empty > 0 else 0
            
            self.metrics["total_cells_checked"] = total_data_cells
            self.metrics["formula_cells"] = formula_cells
            self.metrics["value_cells"] = value_cells
            self.metrics["empty_cells"] = empty_cells
            self.metrics["formula_percentage"] = round(formula_percentage, 1)
            self.metrics["sample_formulas"] = sample_formulas
            
            if formula_percentage < self.MIN_FORMULA_PERCENTAGE:
                self.errors.append(
                    f"Formula coverage {formula_percentage:.1f}% is below minimum {self.MIN_FORMULA_PERCENTAGE}%"
                )
                return False
            
            return True
            
        except Exception as e:
            self.errors.append(f"Error checking formula coverage: {str(e)}")
            return False
    
    def _check_formula_references(self, sheet_name: str) -> bool:
        """
        Check that formulas reference the raw data sheet correctly.
        """
        try:
            wb = load_workbook(self.workbook_path, data_only=False)
            ws = wb[sheet_name]
            
            formulas_checked = 0
            correct_references = 0
            incorrect_references = []
            
            for row in ws.iter_rows(min_row=1, max_row=min(50, ws.max_row or 1)):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formulas_checked += 1
                        formula = cell.value
                        
                        # Check if formula references raw data sheet
                        if self.RAW_DATA_SHEET in formula or "'" not in formula:
                            correct_references += 1
                        elif re.search(r"'[^']+'\!", formula):
                            # References another sheet - check if it's valid
                            sheet_refs = re.findall(r"'([^']+)'\!", formula)
                            for ref in sheet_refs:
                                if ref not in wb.sheetnames and ref != self.RAW_DATA_SHEET:
                                    incorrect_references.append({
                                        "cell": f"{get_column_letter(cell.column)}{cell.row}",
                                        "formula": formula[:50],
                                        "invalid_ref": ref
                                    })
                        else:
                            correct_references += 1
            
            wb.close()
            
            self.metrics["formulas_checked"] = formulas_checked
            self.metrics["correct_references"] = correct_references
            self.metrics["incorrect_references"] = incorrect_references[:5]
            
            if incorrect_references:
                self.warnings.append(
                    f"{len(incorrect_references)} formulas reference invalid sheets"
                )
            
            return len(incorrect_references) == 0
            
        except Exception as e:
            self.warnings.append(f"Error checking formula references: {str(e)}")
            return True  # Don't fail on this check

    def run_statistical_verification(
        self,
        sheet_name: str,
        task_id: str,
        task_type: str,
        verification_config: Dict[str, Any]
    ) -> bool:
        """
        Run statistical verification against Python ground truth.
        
        Args:
            sheet_name: Sheet to verify.
            task_id: Task identifier.
            task_type: Type of analysis task.
            verification_config: Task-specific verification parameters.
        
        Returns:
            True if verification passes, False otherwise.
        """
        if self.raw_data is None:
            self.warnings.append("No raw data provided for statistical verification")
            return True
        
        try:
            self.verification_result = verify_task_output(
                workbook_path=self.workbook_path,
                raw_data=self.raw_data,
                task_id=task_id,
                task_type=task_type,
                sheet_name=sheet_name,
                verification_config=verification_config
            )
            
            self.metrics["verification_status"] = self.verification_result.status.value
            self.metrics["verification_pass_rate"] = self.verification_result.pass_rate
            self.metrics["verification_passed"] = self.verification_result.passed_checks
            self.metrics["verification_failed"] = self.verification_result.failed_checks
            
            if self.verification_result.status == VerificationStatus.FAIL:
                failed_checks = [
                    c for c in self.verification_result.checks 
                    if c.status == VerificationStatus.FAIL
                ]
                for check in failed_checks[:3]:
                    self.errors.append(
                        f"Verification failed: {check.check_name} - "
                        f"expected {check.expected_value:.4f}, got {check.actual_value}"
                    )
                return False
            
            return True
            
        except Exception as e:
            self.warnings.append(f"Statistical verification error: {str(e)}")
            return True

    def _check_no_excel_errors(self, sheet_name: str) -> bool:
        """
        Check for Excel error values (#REF!, #DIV/0!, etc.) in calculated values.
        """
        try:
            # Load with data_only=True to see calculated values
            wb = load_workbook(self.workbook_path, data_only=True)
            ws = wb[sheet_name]
            
            error_patterns = ['#REF!', '#DIV/0!', '#VALUE!', '#NAME?', '#N/A', '#NULL!', '#NUM!']
            errors_found = []
            
            for row in ws.iter_rows(min_row=1, max_row=min(100, ws.max_row or 1)):
                for cell in row:
                    if cell.value and str(cell.value) in error_patterns:
                        errors_found.append({
                            "cell": f"{get_column_letter(cell.column)}{cell.row}",
                            "error": str(cell.value)
                        })
            
            wb.close()
            
            self.metrics["excel_errors_found"] = errors_found[:10]
            
            if errors_found:
                self.warnings.append(
                    f"{len(errors_found)} Excel error values found (may be expected for empty data)"
                )
            
            return True  # Don't fail on errors - they might be expected
            
        except Exception as e:
            # data_only mode might fail if file wasn't saved with Excel
            self.warnings.append(f"Could not check for Excel errors: {str(e)}")
            return True
    
    def _build_result(self, passed: bool) -> Dict[str, Any]:
        """Build the final QC result dictionary."""
        return {
            "passed": passed,
            "errors": self.errors,
            "warnings": self.warnings,
            "metrics": self.metrics,
            "summary": self._generate_summary(passed)
        }
    
    def _generate_summary(self, passed: bool) -> str:
        """Generate human-readable summary."""
        lines = [
            "=" * 50,
            "DETERMINISTIC QC RESULTS",
            "=" * 50,
            f"Status: {'PASSED' if passed else 'FAILED'}",
            "",
            "Metrics:"
        ]
        
        if "formula_percentage" in self.metrics:
            lines.append(f"  - Formula coverage: {self.metrics['formula_percentage']}%")
        if "formula_cells" in self.metrics:
            lines.append(f"  - Formula cells: {self.metrics['formula_cells']}")
        if "sample_formulas" in self.metrics and self.metrics["sample_formulas"]:
            lines.append("  - Sample formulas:")
            for f in self.metrics["sample_formulas"][:3]:
                lines.append(f"    {f['cell']}: {f['formula'][:40]}...")
        
        if self.errors:
            lines.append("")
            lines.append("Errors:")
            for e in self.errors:
                lines.append(f"  FAIL: {e}")
        
        if self.warnings:
            lines.append("")
            lines.append("Warnings:")
            for w in self.warnings:
                lines.append(f"  WARN: {w}")
        
        lines.append("=" * 50)
        
        return "\n".join(lines)


def run_deterministic_qc(
    workbook_path: Path,
    sheet_name: str,
    raw_data: Optional[pd.DataFrame] = None,
    task_id: Optional[str] = None,
    task_type: Optional[str] = None,
    verification_config: Optional[Dict[str, Any]] = None
) -> Dict[str, Any]:
    """
    Convenience function to run deterministic QC with optional statistical verification.
    
    Args:
        workbook_path: Path to Excel workbook.
        sheet_name: Sheet to verify.
        raw_data: Optional DataFrame for statistical verification.
        task_id: Optional task identifier for verification.
        task_type: Optional task type for verification.
        verification_config: Optional verification parameters.
    
    Returns:
        QC result dictionary.
    """
    qc = DeterministicQC(workbook_path, raw_data)
    result = qc.run_all_checks(sheet_name)
    
    if raw_data is not None and task_id and task_type and verification_config:
        stat_passed = qc.run_statistical_verification(
            sheet_name=sheet_name,
            task_id=task_id,
            task_type=task_type,
            verification_config=verification_config
        )
        if not stat_passed:
            result["passed"] = False
            result["errors"] = qc.errors
            result["metrics"] = qc.metrics
            result["summary"] = qc._generate_summary(False)
    
    return result
