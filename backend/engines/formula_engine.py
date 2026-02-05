"""
Deterministic Formula Engine.
Generates Excel formulas programmatically by task type.
No LLM involvement - pure template-based generation.
"""

from typing import Dict, List, Any, Tuple, Optional
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, PatternFill

from models.task_schema import TaskType, TaskSpec
from tools.excel_template import ensure_macro_workbook


class FormulaEngine:
    """
    Deterministic formula generation engine.
    Each task type has a dedicated method that generates formulas programmatically.
    """

    def __init__(self, workbook_path: Path, df: pd.DataFrame, session_id: str):
        self.workbook_path = workbook_path
        self.df = df
        self.session_id = session_id
        self.n_rows = len(df)
        self.raw_sheet = "00_RAW_DATA_LOCKED"
        self.clean_sheet = "00_CLEANED_DATA"
        self.normalized_sheet = "00_NORMALIZED_DATA"
        self.data_sheet = self.raw_sheet

        # Build column mapping: column_name -> Excel letter
        self.col_mapping: Dict[str, str] = {}
        for i, col in enumerate(df.columns):
            self.col_mapping[col] = get_column_letter(i + 1)

        # Identify column types
        self.cleaned_df, self.numeric_cols, self.categorical_cols = self._clean_dataframe(df)

        # Styles
        self.header_font = Font(bold=True)
        self.header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _clean_dataframe(self, df: pd.DataFrame) -> Tuple[pd.DataFrame, List[str], List[str]]:
        """
        Clean dataframe for analysis and detect numeric vs categorical columns.
        Numeric detection attempts coercion for object columns.
        """
        cleaned = df.copy()
        numeric_cols: List[str] = []
        categorical_cols: List[str] = []

        for col in df.columns:
            series = df[col]
            if pd.api.types.is_numeric_dtype(series):
                cleaned[col] = pd.to_numeric(series, errors="coerce")
                numeric_cols.append(col)
                continue

            numeric_candidate = pd.to_numeric(series, errors="coerce")
            non_null = series.notna().sum()
            numeric_ratio = numeric_candidate.notna().sum() / max(non_null, 1)

            if non_null >= 5 and numeric_ratio >= 0.8:
                cleaned[col] = numeric_candidate
                numeric_cols.append(col)
            else:
                clean_series = series.astype(str).str.strip()
                clean_series = clean_series.replace({
                    "": pd.NA,
                    "nan": pd.NA,
                    "NaN": pd.NA,
                    "None": pd.NA
                })
                cleaned[col] = clean_series
                categorical_cols.append(col)

        return cleaned, numeric_cols, categorical_cols

    def _format_criteria(self, value: Any) -> str:
        """Format Excel criteria for COUNTIF/COUNTIFS based on value type."""
        try:
            if pd.isna(value):
                return "\"\""
        except Exception:
            pass

        if isinstance(value, bool):
            return "TRUE" if value else "FALSE"
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return str(value)

        text = str(value).replace('"', '""')
        return f"\"{text}\""

    def _row_count_formula(self, sheet_name: str) -> str:
        """Return a formula to estimate total data rows across all columns."""
        letters = list(self.col_mapping.values())
        if not letters:
            return "=0"
        parts = [f"COUNTA('{sheet_name}'!{letter}:{letter})" for letter in letters]
        return f"=MAX({','.join(parts)})-1"

    def _select_data_sheet(self, wb: Workbook) -> None:
        """Choose the analysis data sheet (cleaned preferred)."""
        if self.clean_sheet in wb.sheetnames:
            self.data_sheet = self.clean_sheet
        else:
            self.data_sheet = self.raw_sheet

    def _ensure_raw_data_sheet(self, wb: Workbook) -> None:
        """Ensure raw data sheet exists and is populated with values."""
        if self.raw_sheet in wb.sheetnames:
            return

        ws = wb.create_sheet(self.raw_sheet, 0)
        for i, col in enumerate(self.df.columns, 1):
            ws.cell(row=1, column=i, value=col)

        for row_idx, row in enumerate(self.df.values, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        ws.freeze_panes = "A2"
        ws.protection.sheet = True
        ws.protection.password = "locked"

    def _ensure_cleaned_sheet(self, wb: Workbook) -> None:
        """Create cleaned data sheet with formulas if missing."""
        if self.clean_sheet in wb.sheetnames:
            return

        ws = wb.create_sheet(self.clean_sheet, 1)
        for i, col in enumerate(self.df.columns, 1):
            ws.cell(row=1, column=i, value=col)

        for row_idx in range(2, self.n_rows + 2):
            for col_idx, col_name in enumerate(self.df.columns, 1):
                col_letter = self.col_mapping[col_name]
                raw_cell = f"'{self.raw_sheet}'!{col_letter}{row_idx}"
                if col_name in self.numeric_cols:
                    formula = f'=IF({raw_cell}="","",IFERROR(VALUE({raw_cell}),""))'
                else:
                    formula = f'=IF({raw_cell}="","",TRIM({raw_cell}))'
                ws.cell(row=row_idx, column=col_idx, value=formula)

        ws.freeze_panes = "A2"

    def _ensure_normalized_sheet(self, wb: Workbook) -> None:
        """Create z-score normalized sheet for numeric columns if missing."""
        if self.normalized_sheet in wb.sheetnames:
            return

        ws = wb.create_sheet(self.normalized_sheet, 2)
        for i, col in enumerate(self.df.columns, 1):
            ws.cell(row=1, column=i, value=col)

        for row_idx in range(2, self.n_rows + 2):
            for col_idx, col_name in enumerate(self.df.columns, 1):
                col_letter = self.col_mapping[col_name]
                clean_cell = f"'{self.clean_sheet}'!{col_letter}{row_idx}"
                data_range = f"'{self.clean_sheet}'!{col_letter}2:{col_letter}{self.n_rows + 1}"
                if col_name in self.numeric_cols:
                    formula = (
                        f'=IF({clean_cell}="","",'
                        f'IFERROR(({clean_cell}-AVERAGE({data_range}))/STDEV.S({data_range}),""))'
                    )
                else:
                    formula = f"={clean_cell}"
                ws.cell(row=row_idx, column=col_idx, value=formula)

        ws.freeze_panes = "A2"

    def _write_group_helpers(
        self,
        ws: Worksheet,
        group_var: str,
        cols_to_use: List[str],
        start_col: int = 14,
        header_row: int = 1
    ) -> Tuple[Any, Any, Dict[str, Tuple[str, str]]]:
        """
        Write hidden helper columns for two-group comparisons.
        Returns group1, group2, and a map of column -> (range1, range2).
        """
        groups = self.cleaned_df[group_var].dropna().unique()
        if len(groups) < 2:
            raise ValueError("Need at least 2 groups for comparison")

        group1, group2 = groups[0], groups[1]
        criteria1 = self._format_criteria(group1)
        criteria2 = self._format_criteria(group2)
        group_letter = self.col_mapping[group_var]

        helper_ranges: Dict[str, Tuple[str, str]] = {}

        for idx, col_name in enumerate(cols_to_use):
            if col_name not in self.col_mapping:
                continue
            data_letter = self.col_mapping[col_name]
            col1_idx = start_col + (idx * 2)
            col2_idx = col1_idx + 1
            letter1 = get_column_letter(col1_idx)
            letter2 = get_column_letter(col2_idx)

            ws.cell(row=header_row, column=col1_idx, value=f"{col_name}_G1")
            ws.cell(row=header_row, column=col2_idx, value=f"{col_name}_G2")

            for r in range(2, self.n_rows + 2):
                group_cell = f"'{self.data_sheet}'!{group_letter}{r}"
                data_cell = f"'{self.data_sheet}'!{data_letter}{r}"
                ws.cell(row=r, column=col1_idx, value=f"=IF({group_cell}={criteria1},{data_cell},\"\")")
                ws.cell(row=r, column=col2_idx, value=f"=IF({group_cell}={criteria2},{data_cell},\"\")")

            ws.column_dimensions[letter1].hidden = True
            ws.column_dimensions[letter2].hidden = True

            range1 = f"{letter1}2:{letter1}{self.n_rows + 1}"
            range2 = f"{letter2}2:{letter2}{self.n_rows + 1}"
            helper_ranges[col_name] = (range1, range2)

        return group1, group2, helper_ranges

    def execute_task(self, task: TaskSpec) -> Dict[str, Any]:
        """
        Execute a task and return results.
        Routes to the appropriate task-type method.
        """
        task_methods = {
            TaskType.DATA_AUDIT: self._create_data_audit,
            TaskType.DATA_DICTIONARY: self._create_data_dictionary,
            TaskType.MISSING_DATA: self._create_missing_data,
            TaskType.DESCRIPTIVE_STATS: self._create_descriptive_stats,
            TaskType.FREQUENCY_TABLES: self._create_frequency_tables,
            TaskType.NORMALITY_CHECK: self._create_normality_check,
            TaskType.CORRELATION_MATRIX: self._create_correlation_matrix,
            TaskType.RELIABILITY_ALPHA: self._create_reliability_alpha,
            TaskType.GROUP_COMPARISON: self._create_group_comparison,
            TaskType.CROSS_TABULATION: self._create_cross_tabulation,
            TaskType.EFFECT_SIZES: self._create_effect_sizes,
            TaskType.SUMMARY_DASHBOARD: self._create_summary_dashboard,
        }

        method = task_methods.get(task.task_type)
        if not method:
            raise ValueError(f"Unknown task type: {task.task_type}")

        return method(task)

    def _get_data_range(self, col_name: str, sheet_name: Optional[str] = None) -> str:
        """Get Excel range reference for a column's data."""
        col_letter = self.col_mapping.get(col_name)
        if not col_letter:
            raise ValueError(f"Column '{col_name}' not found")
        sheet = sheet_name or self.data_sheet
        return f"'{sheet}'!{col_letter}2:{col_letter}{self.n_rows + 1}"

    def _open_workbook(self) -> Workbook:
        """Open or create macro-enabled workbook and ensure data sheets exist."""
        if not self.workbook_path.exists():
            ensure_macro_workbook(self.workbook_path)

        wb = load_workbook(self.workbook_path, keep_vba=True)

        self._ensure_raw_data_sheet(wb)
        self._ensure_cleaned_sheet(wb)
        self._ensure_normalized_sheet(wb)
        self._select_data_sheet(wb)

        wb.save(self.workbook_path)
        return wb

    def _create_data_audit(self, task: TaskSpec) -> Dict[str, Any]:
        """Create data audit trail sheet."""
        wb = self._open_workbook()
        ws = wb.create_sheet(task.output_sheet)

        formulas = []

        ws['A1'] = "DATA AUDIT TRAIL"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A3'] = "Session ID:"
        ws['B3'] = self.session_id
        ws['A4'] = "Analysis Date:"
        ws['B4'] = '=TEXT(NOW(),"yyyy-mm-dd hh:mm:ss")'
        formulas.append({"cell": "B4", "formula": ws['B4'].value, "purpose": "Timestamp"})

        ws['A6'] = "DATASET METRICS"
        ws['A6'].font = self.header_font

        data_sheet = self.data_sheet
        row_count_formula = self._row_count_formula(data_sheet)
        col_count_formula = f"=COUNTA('{data_sheet}'!1:1)"

        metrics = [
            ("A7", "Total Rows:", "B7", row_count_formula),
            ("A8", "Total Columns:", "B8", col_count_formula),
            ("A9", "Total Cells:", "B9", f"=B7*B8"),
            ("A10", "Numeric Variables:", "B10", f"={len(self.numeric_cols)}"),
            ("A11", "Categorical Variables:", "B11", f"={len(self.categorical_cols)}"),
        ]

        for label_cell, label, value_cell, formula in metrics:
            ws[label_cell] = label
            ws[value_cell] = formula
            formulas.append({"cell": value_cell, "formula": formula, "purpose": label.replace(":", "")})

        ws['A13'] = "DATA INTEGRITY CHECKS"
        ws['A13'].font = self.header_font

        ws['A14'] = "Total Missing Values:"
        missing_formula = f"=SUMPRODUCT(--(OFFSET('{data_sheet}'!A2,0,0,B7,B8)=\"\"))"
        ws['B14'] = missing_formula
        formulas.append({"cell": "B14", "formula": ws['B14'].value, "purpose": "Missing count"})

        ws['A15'] = "Overall Completeness %:"
        ws['B15'] = f"=ROUND((1-B14/(B7*B8))*100,1)"
        formulas.append({"cell": "B15", "formula": ws['B15'].value, "purpose": "Completeness"})

        wb.save(self.workbook_path)
        wb.close()

        return {
            "sheet_name": task.output_sheet,
            "formulas_created": len(formulas),
            "formulas": formulas
        }

    def _create_data_dictionary(self, task: TaskSpec) -> Dict[str, Any]:
        """Create comprehensive data dictionary."""
        wb = self._open_workbook()
        ws = wb.create_sheet(task.output_sheet)

        formulas = []

        ws['A1'] = "DATA DICTIONARY"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        headers = ["Variable", "Column", "Type", "Level", "N Valid", "N Missing",
                   "% Complete", "Min", "Max", "Mean/Mode", "SD", "Unique"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=i, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill

        row = 5
        for col_name in self.df.columns:
            col_letter = self.col_mapping.get(col_name)
            if not col_letter:
                continue

            data_range = self._get_data_range(col_name)
            is_numeric = col_name in self.numeric_cols
            col_data = self.cleaned_df[col_name]
            unique_count = col_data.nunique(dropna=True)

            if is_numeric:
                if unique_count <= 2:
                    var_type, meas_level = "Binary", "Nominal"
                elif unique_count <= 7:
                    var_type, meas_level = "Ordinal", "Ordinal"
                else:
                    var_type, meas_level = "Continuous", "Interval/Ratio"
            else:
                var_type, meas_level = "Categorical", "Nominal"

            ws.cell(row=row, column=1, value=col_name)
            ws.cell(row=row, column=2, value=col_letter)
            ws.cell(row=row, column=3, value=var_type)
            ws.cell(row=row, column=4, value=meas_level)

            f_valid = f"=COUNTA({data_range})"
            f_missing = f"=COUNTBLANK({data_range})"
            f_complete = f"=ROUND(COUNTA({data_range})/ROWS({data_range})*100,1)"
            f_min = f"=IFERROR(MIN({data_range}),\"-\")"
            f_max = f"=IFERROR(MAX({data_range}),\"-\")"
            f_central = f"=IFERROR(ROUND(AVERAGE({data_range}),2),\"-\")"
            f_sd = f"=IFERROR(ROUND(STDEV.S({data_range}),2),\"-\")"
            f_unique = f"=SUMPRODUCT(1/COUNTIFS({data_range},\"<>\"&\"\",{data_range},{data_range}))"

            ws.cell(row=row, column=5, value=f_valid)
            ws.cell(row=row, column=6, value=f_missing)
            ws.cell(row=row, column=7, value=f_complete)
            ws.cell(row=row, column=8, value=f_min)
            ws.cell(row=row, column=9, value=f_max)
            ws.cell(row=row, column=10, value=f_central)
            ws.cell(row=row, column=11, value=f_sd)
            ws.cell(row=row, column=12, value=f_unique)

            formulas.extend([
                {"cell": f"E{row}", "formula": f_valid, "purpose": f"{col_name} N valid"},
                {"cell": f"F{row}", "formula": f_missing, "purpose": f"{col_name} N missing"},
                {"cell": f"G{row}", "formula": f_complete, "purpose": f"{col_name} % complete"},
            ])

            row += 1

        wb.save(self.workbook_path)
        wb.close()

        return {
            "sheet_name": task.output_sheet,
            "formulas_created": len(formulas),
            "formulas": formulas
        }

    def _create_missing_data(self, task: TaskSpec) -> Dict[str, Any]:
        """Create missing data analysis sheet."""
        wb = self._open_workbook()
        ws = wb.create_sheet(task.output_sheet)

        formulas = []

        ws['A1'] = "MISSING DATA ANALYSIS"
        ws['A1'].font = Font(bold=True, size=14)

        headers = ["Variable", "N Total", "N Missing", "% Missing", "Pattern"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=i, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill

        row = 4
        for col_name in self.df.columns:
            data_range = self._get_data_range(col_name)

            ws.cell(row=row, column=1, value=col_name)

            f_total = f"=ROWS({data_range})"
            f_missing = f"=COUNTBLANK({data_range})"
            f_pct = f"=ROUND(COUNTBLANK({data_range})/ROWS({data_range})*100,1)"
            f_pattern = (
                f'=IF(COUNTBLANK({data_range})=0,"Complete",'
                f'IF(COUNTBLANK({data_range})<{self.n_rows}*0.05,"<5%",'
                f'IF(COUNTBLANK({data_range})<{self.n_rows}*0.2,"5-20%",">20%")))'
            )

            ws.cell(row=row, column=2, value=f_total)
            ws.cell(row=row, column=3, value=f_missing)
            ws.cell(row=row, column=4, value=f_pct)
            ws.cell(row=row, column=5, value=f_pattern)

            formulas.extend([
                {"cell": f"C{row}", "formula": f_missing, "purpose": f"{col_name} missing"},
                {"cell": f"D{row}", "formula": f_pct, "purpose": f"{col_name} % missing"},
            ])

            row += 1

        wb.save(self.workbook_path)
        wb.close()

        return {
            "sheet_name": task.output_sheet,
            "formulas_created": len(formulas),
            "formulas": formulas
        }

    def _create_descriptive_stats(self, task: TaskSpec) -> Dict[str, Any]:
        """Create descriptive statistics sheet."""
        wb = self._open_workbook()
        ws = wb.create_sheet(task.output_sheet)

        formulas = []

        ws['A1'] = "DESCRIPTIVE STATISTICS"
        ws['A1'].font = Font(bold=True, size=14)

        headers = ["Variable", "N", "Mean", "SD", "SE", "Median", "Min", "Max", "Range", "Skewness", "Kurtosis"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=i, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill

        cols_to_use = task.columns.column_names if task.columns.column_names else self.numeric_cols
        if task.columns.max_columns:
            cols_to_use = cols_to_use[:task.columns.max_columns]

        row = 4
        for col_name in cols_to_use:
            if col_name not in self.numeric_cols:
                continue

            data_range = self._get_data_range(col_name)

            ws.cell(row=row, column=1, value=col_name)

            stats_formulas = [
                (2, f"=COUNT({data_range})", "N"),
                (3, f"=ROUND(AVERAGE({data_range}),3)", "Mean"),
                (4, f"=ROUND(STDEV.S({data_range}),3)", "SD"),
                (5, f"=ROUND(STDEV.S({data_range})/SQRT(COUNT({data_range})),4)", "SE"),
                (6, f"=ROUND(MEDIAN({data_range}),3)", "Median"),
                (7, f"=MIN({data_range})", "Min"),
                (8, f"=MAX({data_range})", "Max"),
                (9, f"=MAX({data_range})-MIN({data_range})", "Range"),
                (10, f"=ROUND(SKEW({data_range}),3)", "Skewness"),
                (11, f"=ROUND(KURT({data_range}),3)", "Kurtosis"),
            ]

            for col_idx, formula, purpose in stats_formulas:
                ws.cell(row=row, column=col_idx, value=formula)
                formulas.append({
                    "cell": f"{get_column_letter(col_idx)}{row}",
                    "formula": formula,
                    "purpose": f"{col_name} {purpose}"
                })

            row += 1

        wb.save(self.workbook_path)
        wb.close()

        return {
            "sheet_name": task.output_sheet,
            "formulas_created": len(formulas),
            "formulas": formulas
        }

    def _create_frequency_tables(self, task: TaskSpec) -> Dict[str, Any]:
        """Create frequency tables for categorical variables."""
        wb = self._open_workbook()
        ws = wb.create_sheet(task.output_sheet)

        formulas = []

        ws['A1'] = "FREQUENCY TABLES"
        ws['A1'].font = Font(bold=True, size=14)

        cols_to_use = task.columns.column_names if task.columns.column_names else self.categorical_cols
        if task.columns.max_columns:
            cols_to_use = cols_to_use[:task.columns.max_columns]

        current_row = 3
        for col_name in cols_to_use:
            data_range = self._get_data_range(col_name)

            ws.cell(row=current_row, column=1, value=f"Variable: {col_name}")
            ws.cell(row=current_row, column=1).font = self.header_font

            current_row += 1
            headers = ["Value", "Frequency", "Percent", "Cumulative %"]
            for i, h in enumerate(headers, 1):
                ws.cell(row=current_row, column=i, value=h)
                ws.cell(row=current_row, column=i).font = self.header_font

            current_row += 1
            unique_values = self.cleaned_df[col_name].dropna().unique()
            value_start_row = current_row

            for val in unique_values:
                ws.cell(row=current_row, column=1, value=val)
                criteria = self._format_criteria(val)

                f_freq = f"=COUNTIF({data_range},{criteria})"
                f_pct = f"=ROUND(COUNTIF({data_range},{criteria})/COUNTA({data_range})*100,1)"
                f_cum = f"=SUM(C{value_start_row}:C{current_row})"

                ws.cell(row=current_row, column=2, value=f_freq)
                ws.cell(row=current_row, column=3, value=f_pct)
                ws.cell(row=current_row, column=4, value=f_cum)

                formulas.append({"cell": f"B{current_row}", "formula": f_freq, "purpose": f"{col_name}={val} freq"})

                current_row += 1

            current_row += 2

        wb.save(self.workbook_path)
        wb.close()

        return {
            "sheet_name": task.output_sheet,
            "formulas_created": len(formulas),
            "formulas": formulas
        }

    def _create_normality_check(self, task: TaskSpec) -> Dict[str, Any]:
        """Create normality diagnostics using UDFs where available."""
        wb = self._open_workbook()
        ws = wb.create_sheet(task.output_sheet)

        formulas = []

        ws['A1'] = "NORMALITY DIAGNOSTICS"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = "Note: Shapiro-Wilk uses UDF SHAPIRO_WILK(). Skew/Kurt provided as supplemental."
        ws['A2'].font = Font(italic=True)

        headers = ["Variable", "N", "Shapiro W", "Shapiro p", "Skewness", "Kurtosis", "Z Skew", "Z Kurt", "Assessment"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=i, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill

        cols_to_use = task.columns.column_names if task.columns.column_names else self.numeric_cols
        if task.columns.max_columns:
            cols_to_use = cols_to_use[:task.columns.max_columns]

        row = 5
        for col_name in cols_to_use:
            if col_name not in self.numeric_cols:
                continue

            data_range = self._get_data_range(col_name)

            ws.cell(row=row, column=1, value=col_name)

            ws.cell(row=row, column=2, value=f"=COUNT({data_range})")
            ws.cell(row=row, column=3, value=f"=INDEX(SHAPIRO_WILK({data_range}),1)")
            ws.cell(row=row, column=4, value=f"=INDEX(SHAPIRO_WILK({data_range}),2)")
            ws.cell(row=row, column=5, value=f"=ROUND(SKEW({data_range}),3)")
            ws.cell(row=row, column=6, value=f"=ROUND(KURT({data_range}),3)")
            ws.cell(row=row, column=7, value=f"=ROUND(SKEW({data_range})/SQRT(6/COUNT({data_range})),2)")
            ws.cell(row=row, column=8, value=f"=ROUND(KURT({data_range})/SQRT(24/COUNT({data_range})),2)")
            ws.cell(row=row, column=9, value=f'=IF(D{row}>=0.05,"Normal","Non-normal")')

            formulas.extend([
                {"cell": f"C{row}", "formula": f"=INDEX(SHAPIRO_WILK({data_range}),1)", "purpose": f"{col_name} Shapiro W"},
                {"cell": f"D{row}", "formula": f"=INDEX(SHAPIRO_WILK({data_range}),2)", "purpose": f"{col_name} Shapiro p"},
                {"cell": f"E{row}", "formula": f"=SKEW({data_range})", "purpose": f"{col_name} skewness"},
                {"cell": f"F{row}", "formula": f"=KURT({data_range})", "purpose": f"{col_name} kurtosis"},
            ])

            row += 1

        wb.save(self.workbook_path)
        wb.close()

        return {
            "sheet_name": task.output_sheet,
            "formulas_created": len(formulas),
            "formulas": formulas
        }

    def _create_correlation_matrix(self, task: TaskSpec) -> Dict[str, Any]:
        """Create correlation matrix."""
        wb = self._open_workbook()
        ws = wb.create_sheet(task.output_sheet)

        formulas = []

        ws['A1'] = "CORRELATION MATRIX (Pearson r)"
        ws['A1'].font = Font(bold=True, size=14)

        cols_to_use = task.columns.column_names if task.columns.column_names else self.numeric_cols
        if task.columns.max_columns:
            cols_to_use = cols_to_use[:task.columns.max_columns]
        cols_to_use = [c for c in cols_to_use if c in self.numeric_cols]

        for i, col in enumerate(cols_to_use, 2):
            ws.cell(row=3, column=i, value=col[:10])
            ws.cell(row=3, column=i).font = self.header_font

        for i, row_col in enumerate(cols_to_use):
            row = i + 4
            ws.cell(row=row, column=1, value=row_col[:15])
            ws.cell(row=row, column=1).font = self.header_font

            for j, col_col in enumerate(cols_to_use):
                col = j + 2

                if i == j:
                    ws.cell(row=row, column=col, value="=1")
                    formulas.append({"cell": f"{get_column_letter(col)}{row}", "formula": "=1", "purpose": "Diagonal"})
                elif i < j:
                    range1 = self._get_data_range(row_col)
                    range2 = self._get_data_range(col_col)
                    formula = f"=ROUND(CORREL({range1},{range2}),3)"
                    ws.cell(row=row, column=col, value=formula)
                    formulas.append({"cell": f"{get_column_letter(col)}{row}", "formula": formula, "purpose": f"r({row_col},{col_col})"})
                else:
                    ref_row = j + 4
                    ref_col = i + 2
                    formula = f"={get_column_letter(ref_col)}{ref_row}"
                    ws.cell(row=row, column=col, value=formula)

        wb.save(self.workbook_path)
        wb.close()

        return {
            "sheet_name": task.output_sheet,
            "formulas_created": len(formulas),
            "formulas": formulas
        }

    def _create_reliability_alpha(self, task: TaskSpec) -> Dict[str, Any]:
        """Create Cronbach's alpha calculation sheet."""
        wb = self._open_workbook()
        ws = wb.create_sheet(task.output_sheet)

        formulas = []

        ws['A1'] = "RELIABILITY ANALYSIS (Cronbach's Alpha)"
        ws['A1'].font = Font(bold=True, size=14)

        items = task.scale_items if task.scale_items else self.numeric_cols
        items = [i for i in items if i in self.col_mapping]
        k = len(items)

        if k < 2:
            ws['A3'] = "Error: Need at least 2 items for reliability analysis"
            wb.save(self.workbook_path)
            wb.close()
            return {"sheet_name": task.output_sheet, "formulas_created": 0, "formulas": [], "error": "Insufficient items"}

        ws['A3'] = f"Scale: {task.name}"
        ws['A4'] = f"Number of items (k): {k}"

        # Helper matrix for contiguous range (hidden)
        helper_start_col = 8  # Column H
        helper_header_row = 1
        helper_data_start = 2
        helper_data_end = self.n_rows + 1

        item_col_map: Dict[str, str] = {}
        for idx, item in enumerate(items):
            col_idx = helper_start_col + idx
            col_letter = get_column_letter(col_idx)
            item_col_map[item] = col_letter
            ws.cell(row=helper_header_row, column=col_idx, value=f"{item}_CLEAN")

            source_letter = self.col_mapping[item]
            for r in range(helper_data_start, helper_data_end + 1):
                source_cell = f"'{self.data_sheet}'!{source_letter}{r}"
                ws.cell(row=r, column=col_idx, value=f"={source_cell}")

            ws.column_dimensions[col_letter].hidden = True

        total_col_idx = helper_start_col + k
        total_col_letter = get_column_letter(total_col_idx)
        ws.cell(row=helper_header_row, column=total_col_idx, value="TOTAL_SCORE")
        for r in range(helper_data_start, helper_data_end + 1):
            ws.cell(row=r, column=total_col_idx, value=f"=SUM({item_col_map[items[0]]}{r}:{item_col_map[items[-1]]}{r})")
        ws.column_dimensions[total_col_letter].hidden = True

        total_minus_map: Dict[str, str] = {}
        for idx, item in enumerate(items):
            minus_col_idx = helper_start_col + k + 1 + idx
            minus_col_letter = get_column_letter(minus_col_idx)
            total_minus_map[item] = minus_col_letter
            ws.cell(row=helper_header_row, column=minus_col_idx, value=f"TOTAL_MINUS_{item}")
            item_col_letter = item_col_map[item]
            for r in range(helper_data_start, helper_data_end + 1):
                ws.cell(row=r, column=minus_col_idx, value=f"={total_col_letter}{r}-{item_col_letter}{r}")
            ws.column_dimensions[minus_col_letter].hidden = True

        ws['A6'] = "ITEM STATISTICS"
        ws['A6'].font = self.header_font

        headers = ["Item", "Mean", "SD", "Variance", "Item-Total r"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=7, column=i, value=h)
            ws.cell(row=7, column=i).font = self.header_font

        row = 8
        variance_cells = []
        for item in items:
            data_range = self._get_data_range(item)
            item_col_letter = item_col_map[item]
            minus_col_letter = total_minus_map[item]

            ws.cell(row=row, column=1, value=item)
            ws.cell(row=row, column=2, value=f"=ROUND(AVERAGE({data_range}),3)")
            ws.cell(row=row, column=3, value=f"=ROUND(STDEV.S({data_range}),3)")
            var_cell = f"D{row}"
            ws.cell(row=row, column=4, value=f"=ROUND(VAR.S({data_range}),3)")

            item_range = f"{item_col_letter}{helper_data_start}:{item_col_letter}{helper_data_end}"
            minus_range = f"{minus_col_letter}{helper_data_start}:{minus_col_letter}{helper_data_end}"
            item_total_formula = f"=IFERROR(ROUND(CORREL({item_range},{minus_range}),3),\"\")"
            ws.cell(row=row, column=5, value=item_total_formula)

            variance_cells.append(var_cell)
            formulas.append({"cell": f"D{row}", "formula": f"=VAR.S({data_range})", "purpose": f"{item} variance"})
            formulas.append({"cell": f"E{row}", "formula": item_total_formula, "purpose": f"{item} item-total r"})
            row += 1

        alpha_row = row + 2
        ws.cell(row=alpha_row, column=1, value="CRONBACH'S ALPHA")
        ws.cell(row=alpha_row, column=1).font = self.header_font

        ws.cell(row=alpha_row+1, column=1, value="Sum of item variances:")
        sum_var_formula = f"=SUM({variance_cells[0]}:{variance_cells[-1]})"
        ws.cell(row=alpha_row+1, column=2, value=sum_var_formula)

        first_col = item_col_map[items[0]]
        last_col = item_col_map[items[-1]]
        full_range = f"{first_col}{helper_data_start}:{last_col}{helper_data_end}"

        ws.cell(row=alpha_row+2, column=1, value="Cronbach's Alpha (UDF):")
        alpha_udf = f"=ROUND(CRONBACH_ALPHA({full_range}),3)"
        ws.cell(row=alpha_row+2, column=2, value=alpha_udf)
        formulas.append({"cell": f"B{alpha_row+2}", "formula": alpha_udf, "purpose": "Cronbach's Alpha (UDF)"})

        ws.cell(row=alpha_row+4, column=1, value="Interpretation:")
        ws.cell(row=alpha_row+4, column=2, value=f'=IF(B{alpha_row+2}>=0.9,"Excellent",IF(B{alpha_row+2}>=0.8,"Good",IF(B{alpha_row+2}>=0.7,"Acceptable",IF(B{alpha_row+2}>=0.6,"Questionable","Poor"))))')

        wb.save(self.workbook_path)
        wb.close()

        return {
            "sheet_name": task.output_sheet,
            "formulas_created": len(formulas),
            "formulas": formulas
        }

    def _create_group_comparison(self, task: TaskSpec) -> Dict[str, Any]:
        """Create group comparison sheet (t-test style)."""
        wb = self._open_workbook()
        ws = wb.create_sheet(task.output_sheet)

        formulas = []

        ws['A1'] = "GROUP COMPARISON ANALYSIS"
        ws['A1'].font = Font(bold=True, size=14)

        group_var = task.group_by
        if not group_var or group_var not in self.df.columns:
            ws['A3'] = "Error: No valid grouping variable specified"
            wb.save(self.workbook_path)
            wb.close()
            return {"sheet_name": task.output_sheet, "formulas_created": 0, "formulas": []}

        cols_to_use = task.columns.column_names if task.columns.column_names else self.numeric_cols
        if task.columns.max_columns:
            cols_to_use = cols_to_use[:task.columns.max_columns]
        cols_to_use = [c for c in cols_to_use if c in self.numeric_cols and c != group_var]

        try:
            group1, group2, helper_ranges = self._write_group_helpers(
                ws, group_var, cols_to_use, start_col=14, header_row=1
            )
        except Exception:
            ws['A3'] = "Error: Need at least 2 groups for comparison"
            wb.save(self.workbook_path)
            wb.close()
            return {"sheet_name": task.output_sheet, "formulas_created": 0, "formulas": []}

        ws['A3'] = f"Grouping Variable: {group_var}"
        ws['A4'] = f"Group 1: {group1}"
        ws['A5'] = f"Group 2: {group2}"

        headers = ["Variable", "N1", "M1", "SD1", "N2", "M2", "SD2", "Mean Diff", "t", "df", "p", "Cohen's d", "Levene p"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=7, column=i, value=h)
            ws.cell(row=7, column=i).font = self.header_font

        row = 8
        for col_name in cols_to_use:
            if col_name not in helper_ranges:
                continue

            range1, range2 = helper_ranges[col_name]

            ws.cell(row=row, column=1, value=col_name)

            n1_formula = f"=COUNT({range1})"
            n2_formula = f"=COUNT({range2})"
            m1_formula = f"=IFERROR(ROUND(AVERAGE({range1}),3),\"\")"
            m2_formula = f"=IFERROR(ROUND(AVERAGE({range2}),3),\"\")"
            sd1_formula = f"=IFERROR(ROUND(STDEV.S({range1}),3),\"\")"
            sd2_formula = f"=IFERROR(ROUND(STDEV.S({range2}),3),\"\")"

            ws.cell(row=row, column=2, value=n1_formula)
            ws.cell(row=row, column=3, value=m1_formula)
            ws.cell(row=row, column=4, value=sd1_formula)
            ws.cell(row=row, column=5, value=n2_formula)
            ws.cell(row=row, column=6, value=m2_formula)
            ws.cell(row=row, column=7, value=sd2_formula)

            ws.cell(row=row, column=8, value=f"=IF(OR(B{row}=\"\",E{row}=\"\"),\"\",ROUND(C{row}-F{row},3))")

            t_formula = (
                f"=IF(OR(B{row}<2,E{row}<2),\"\","
                f"(C{row}-F{row})/(SQRT(((B{row}-1)*D{row}^2+(E{row}-1)*G{row}^2)/(B{row}+E{row}-2))*SQRT(1/B{row}+1/E{row})))"
                f")"
            )
            df_formula = f"=IF(OR(B{row}<2,E{row}<2),\"\",B{row}+E{row}-2)"
            p_formula = f"=IF(J{row}=\"\",\"\",P_VALUE_T(I{row},J{row}))"
            d_formula = f"=IF(OR(B{row}<2,E{row}<2),\"\",COHENS_D(C{row},D{row},B{row},F{row},G{row},E{row}))"
            levene_formula = f"=IFERROR(INDEX(LEVENE_TEST({range1},{range2}),2),\"\")"

            ws.cell(row=row, column=9, value=t_formula)
            ws.cell(row=row, column=10, value=df_formula)
            ws.cell(row=row, column=11, value=p_formula)
            ws.cell(row=row, column=12, value=d_formula)
            ws.cell(row=row, column=13, value=levene_formula)

            formulas.extend([
                {"cell": f"B{row}", "formula": n1_formula, "purpose": f"{col_name} N1"},
                {"cell": f"E{row}", "formula": n2_formula, "purpose": f"{col_name} N2"},
                {"cell": f"C{row}", "formula": m1_formula, "purpose": f"{col_name} M1"},
                {"cell": f"F{row}", "formula": m2_formula, "purpose": f"{col_name} M2"},
                {"cell": f"I{row}", "formula": t_formula, "purpose": f"{col_name} t"},
                {"cell": f"K{row}", "formula": p_formula, "purpose": f"{col_name} p"},
                {"cell": f"L{row}", "formula": d_formula, "purpose": f"{col_name} Cohen d"},
                {"cell": f"M{row}", "formula": levene_formula, "purpose": f"{col_name} Levene p"},
            ])

            row += 1

        wb.save(self.workbook_path)
        wb.close()

        return {
            "sheet_name": task.output_sheet,
            "formulas_created": len(formulas),
            "formulas": formulas
        }

    def _create_cross_tabulation(self, task: TaskSpec) -> Dict[str, Any]:
        """Create cross-tabulation sheet."""
        wb = self._open_workbook()
        ws = wb.create_sheet(task.output_sheet)

        formulas = []

        ws['A1'] = "CROSS-TABULATION"
        ws['A1'].font = Font(bold=True, size=14)

        cols = task.columns.column_names if task.columns.column_names else self.categorical_cols
        row_var = None
        col_var = None

        if cols and len(cols) >= 2:
            row_var, col_var = cols[0], cols[1]
        elif cols and len(cols) == 1 and task.group_by:
            row_var, col_var = cols[0], task.group_by
        elif task.group_by and task.group_by in self.df.columns:
            row_var = task.group_by
            col_var = next((c for c in self.categorical_cols if c != row_var), None)
        elif len(self.categorical_cols) >= 2:
            row_var, col_var = self.categorical_cols[0], self.categorical_cols[1]

        if not row_var or not col_var:
            ws['A3'] = "Error: Need two categorical variables for cross-tabulation"
            wb.save(self.workbook_path)
            wb.close()
            return {"sheet_name": task.output_sheet, "formulas_created": 0, "formulas": []}

        row_levels = list(self.cleaned_df[row_var].dropna().unique())
        col_levels = list(self.cleaned_df[col_var].dropna().unique())

        if len(row_levels) < 2 or len(col_levels) < 2:
            ws['A5'] = "Error: Need at least 2 categories in each variable"
            wb.save(self.workbook_path)
            wb.close()
            return {"sheet_name": task.output_sheet, "formulas_created": 0, "formulas": []}

        ws['A3'] = f"Row Variable: {row_var}"
        ws['A4'] = f"Column Variable: {col_var}"

        row_range = self._get_data_range(row_var)
        col_range = self._get_data_range(col_var)

        header_row = 6
        ws.cell(row=header_row, column=1, value=row_var)
        for j, level in enumerate(col_levels, 0):
            ws.cell(row=header_row, column=2 + j, value=level)
        row_total_col = 2 + len(col_levels)
        ws.cell(row=header_row, column=row_total_col, value="Row Total")

        data_start_row = header_row + 1
        for i, r_level in enumerate(row_levels):
            row = data_start_row + i
            ws.cell(row=row, column=1, value=r_level)
            for j, c_level in enumerate(col_levels):
                col = 2 + j
                r_crit = self._format_criteria(r_level)
                c_crit = self._format_criteria(c_level)
                formula = f"=COUNTIFS({row_range},{r_crit},{col_range},{c_crit})"
                ws.cell(row=row, column=col, value=formula)
                formulas.append({
                    "cell": f"{get_column_letter(col)}{row}",
                    "formula": formula,
                    "purpose": f"{row_var}={r_level}, {col_var}={c_level}"
                })
            ws.cell(row=row, column=row_total_col, value=f"=SUM(B{row}:{get_column_letter(row_total_col - 1)}{row})")

        total_row = data_start_row + len(row_levels)
        ws.cell(row=total_row, column=1, value="Column Total")
        for j in range(len(col_levels)):
            col = 2 + j
            col_letter = get_column_letter(col)
            ws.cell(row=total_row, column=col, value=f"=SUM({col_letter}{data_start_row}:{col_letter}{total_row - 1})")
        ws.cell(row=total_row, column=row_total_col, value=f"=SUM(B{total_row}:{get_column_letter(row_total_col - 1)}{total_row})")

        # Expected counts table
        expected_header = total_row + 3
        ws.cell(row=expected_header - 1, column=1, value="EXPECTED COUNTS")
        ws.cell(row=expected_header, column=1, value=row_var)
        for j, level in enumerate(col_levels, 0):
            ws.cell(row=expected_header, column=2 + j, value=level)

        expected_start_row = expected_header + 1
        grand_total_cell = f"{get_column_letter(row_total_col)}{total_row}"
        for i, r_level in enumerate(row_levels):
            row = expected_start_row + i
            ws.cell(row=row, column=1, value=r_level)
            row_total_cell = f"{get_column_letter(row_total_col)}{data_start_row + i}"
            for j in range(len(col_levels)):
                col = 2 + j
                col_total_cell = f"{get_column_letter(col)}{total_row}"
                formula = f"=IFERROR({row_total_cell}*{col_total_cell}/{grand_total_cell},\"\")"
                ws.cell(row=row, column=col, value=formula)

        obs_range = f"{get_column_letter(2)}{data_start_row}:{get_column_letter(row_total_col - 1)}{total_row - 1}"
        exp_range = f"{get_column_letter(2)}{expected_start_row}:{get_column_letter(row_total_col - 1)}{expected_start_row + len(row_levels) - 1}"

        chi_row = expected_start_row + len(row_levels) + 2
        ws.cell(row=chi_row, column=1, value="Chi-square")
        chi_formula = f"=SUMPRODUCT(({obs_range}-{exp_range})^2/{exp_range})"
        ws.cell(row=chi_row, column=2, value=chi_formula)
        formulas.append({"cell": f"B{chi_row}", "formula": chi_formula, "purpose": "Chi-square"})

        df_row = chi_row + 1
        ws.cell(row=df_row, column=1, value="df")
        df_formula = f"={(len(row_levels) - 1)}*{(len(col_levels) - 1)}"
        ws.cell(row=df_row, column=2, value=df_formula)

        p_row = chi_row + 2
        ws.cell(row=p_row, column=1, value="p")
        p_formula = f"=CHISQ.DIST.RT(B{chi_row},B{df_row})"
        ws.cell(row=p_row, column=2, value=p_formula)
        formulas.append({"cell": f"B{p_row}", "formula": p_formula, "purpose": "Chi-square p-value"})

        v_row = chi_row + 3
        ws.cell(row=v_row, column=1, value="Cramer's V")
        min_dim = min(len(row_levels) - 1, len(col_levels) - 1)
        v_formula = f"=CRAMERS_V(B{chi_row},{grand_total_cell},{min_dim})"
        ws.cell(row=v_row, column=2, value=v_formula)
        formulas.append({"cell": f"B{v_row}", "formula": v_formula, "purpose": "Cramer's V"})

        wb.save(self.workbook_path)
        wb.close()

        return {
            "sheet_name": task.output_sheet,
            "formulas_created": len(formulas),
            "formulas": formulas
        }

    def _create_effect_sizes(self, task: TaskSpec) -> Dict[str, Any]:
        """Create effect size calculations sheet."""
        wb = self._open_workbook()
        ws = wb.create_sheet(task.output_sheet)

        formulas = []

        ws['A1'] = "EFFECT SIZE CALCULATIONS"
        ws['A1'].font = Font(bold=True, size=14)

        ws['A3'] = "Cohen's d Interpretation:"
        ws['A4'] = "Small: |d| ~ 0.2"
        ws['A5'] = "Medium: |d| ~ 0.5"
        ws['A6'] = "Large: |d| ~ 0.8"

        cols_to_use = task.columns.column_names if task.columns.column_names else self.numeric_cols
        if task.columns.max_columns:
            cols_to_use = cols_to_use[:task.columns.max_columns]
        cols_to_use = [c for c in cols_to_use if c in self.numeric_cols]

        row_cursor = 8

        # Section A: Group effect sizes (Cohen's d)
        if task.group_by and task.group_by in self.df.columns:
            try:
                group1, group2, helper_ranges = self._write_group_helpers(
                    ws, task.group_by, cols_to_use, start_col=20, header_row=1
                )
                ws.cell(row=row_cursor, column=1, value=f"GROUP EFFECT SIZES (Cohen's d) - {task.group_by}")
                ws.cell(row=row_cursor, column=1).font = self.header_font
                row_cursor += 1

                headers = ["Variable", "N1", "M1", "SD1", "N2", "M2", "SD2", "Cohen's d", "Magnitude"]
                for i, h in enumerate(headers, 1):
                    ws.cell(row=row_cursor, column=i, value=h)
                    ws.cell(row=row_cursor, column=i).font = self.header_font
                row_cursor += 1

                for col_name in cols_to_use:
                    if col_name not in helper_ranges:
                        continue
                    range1, range2 = helper_ranges[col_name]

                    ws.cell(row=row_cursor, column=1, value=col_name)
                    n1_formula = f"=COUNT({range1})"
                    n2_formula = f"=COUNT({range2})"
                    m1_formula = f"=IFERROR(ROUND(AVERAGE({range1}),3),\"\")"
                    m2_formula = f"=IFERROR(ROUND(AVERAGE({range2}),3),\"\")"
                    sd1_formula = f"=IFERROR(ROUND(STDEV.S({range1}),3),\"\")"
                    sd2_formula = f"=IFERROR(ROUND(STDEV.S({range2}),3),\"\")"
                    d_formula = f"=IF(OR(B{row_cursor}<2,E{row_cursor}<2),\"\",COHENS_D(C{row_cursor},D{row_cursor},B{row_cursor},F{row_cursor},G{row_cursor},E{row_cursor}))"
                    mag_formula = (
                        f"=IF(H{row_cursor}=\"\",\"\","
                        f"IF(ABS(H{row_cursor})<0.2,\"Negligible\","
                        f"IF(ABS(H{row_cursor})<0.5,\"Small\","
                        f"IF(ABS(H{row_cursor})<0.8,\"Medium\",\"Large\"))))"
                    )

                    ws.cell(row=row_cursor, column=2, value=n1_formula)
                    ws.cell(row=row_cursor, column=3, value=m1_formula)
                    ws.cell(row=row_cursor, column=4, value=sd1_formula)
                    ws.cell(row=row_cursor, column=5, value=n2_formula)
                    ws.cell(row=row_cursor, column=6, value=m2_formula)
                    ws.cell(row=row_cursor, column=7, value=sd2_formula)
                    ws.cell(row=row_cursor, column=8, value=d_formula)
                    ws.cell(row=row_cursor, column=9, value=mag_formula)

                    formulas.append({"cell": f"H{row_cursor}", "formula": d_formula, "purpose": f"{col_name} Cohen d"})
                    row_cursor += 1
            except Exception:
                ws.cell(row=row_cursor, column=1, value="Group effect sizes skipped (insufficient groups)")
                row_cursor += 2

        # Section B: Correlation effect sizes (r, r^2, Fisher z)
        ws.cell(row=row_cursor, column=1, value="CORRELATION EFFECT SIZES")
        ws.cell(row=row_cursor, column=1).font = self.header_font
        row_cursor += 1

        headers = ["Var 1", "Var 2", "r", "r^2", "Fisher z", "Magnitude"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=row_cursor, column=i, value=h)
            ws.cell(row=row_cursor, column=i).font = self.header_font
        row_cursor += 1

        for i in range(len(cols_to_use)):
            for j in range(i + 1, len(cols_to_use)):
                col1 = cols_to_use[i]
                col2 = cols_to_use[j]
                range1 = self._get_data_range(col1)
                range2 = self._get_data_range(col2)

                ws.cell(row=row_cursor, column=1, value=col1)
                ws.cell(row=row_cursor, column=2, value=col2)

                r_formula = f"=ROUND(CORREL({range1},{range2}),3)"
                r2_formula = f"=IF(C{row_cursor}=\"\",\"\",ROUND(C{row_cursor}^2,3))"
                z_formula = f"=IF(C{row_cursor}=\"\",\"\",ROUND(FISHER_Z(C{row_cursor}),3))"
                mag_formula = (
                    f"=IF(C{row_cursor}=\"\",\"\","
                    f"IF(ABS(C{row_cursor})<0.1,\"Negligible\","
                    f"IF(ABS(C{row_cursor})<0.3,\"Small\","
                    f"IF(ABS(C{row_cursor})<0.5,\"Medium\",\"Large\"))))"
                )

                ws.cell(row=row_cursor, column=3, value=r_formula)
                ws.cell(row=row_cursor, column=4, value=r2_formula)
                ws.cell(row=row_cursor, column=5, value=z_formula)
                ws.cell(row=row_cursor, column=6, value=mag_formula)

                formulas.append({"cell": f"C{row_cursor}", "formula": r_formula, "purpose": f"r({col1},{col2})"})
                row_cursor += 1

        wb.save(self.workbook_path)
        wb.close()

        return {
            "sheet_name": task.output_sheet,
            "formulas_created": len(formulas),
            "formulas": formulas
        }

    def _create_summary_dashboard(self, task: TaskSpec) -> Dict[str, Any]:
        """Create summary dashboard sheet."""
        wb = self._open_workbook()
        ws = wb.create_sheet(task.output_sheet)

        formulas = []

        ws['A1'] = "ANALYSIS SUMMARY DASHBOARD"
        ws['A1'].font = Font(bold=True, size=14)

        ws['A3'] = "Dataset Overview"
        ws['A3'].font = self.header_font

        ws['A4'] = "Total Observations:"
        ws['B4'] = self._row_count_formula(self.data_sheet)
        ws['A5'] = "Total Variables:"
        ws['B5'] = f"=COUNTA('{self.data_sheet}'!1:1)"
        ws['A6'] = "Numeric Variables:"
        ws['B6'] = f"={len(self.numeric_cols)}"
        ws['A7'] = "Categorical Variables:"
        ws['B7'] = f"={len(self.categorical_cols)}"

        formulas.extend([
            {"cell": "B4", "formula": ws['B4'].value, "purpose": "N observations"},
            {"cell": "B5", "formula": ws['B5'].value, "purpose": "N variables"},
        ])

        wb.save(self.workbook_path)
        wb.close()

        return {
            "sheet_name": task.output_sheet,
            "formulas_created": len(formulas),
            "formulas": formulas
        }

